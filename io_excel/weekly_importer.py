"""Excel Importer for Full-Year Weekly Curriculum (35 weeks).

Supports formats such as:
- 'Định lượng số tiết theo tuần năm học 2026_2027.xlsx'
- Grade-based semester sheets: HKI_K6, HK2_K6, HKI_K7, HK2_K7, HKI_K8, HK2_K8, HKI_K9, HK2_K9
"""
from __future__ import annotations

import io
import re
import sqlite3
from typing import BinaryIO, Optional, Union
import openpyxl

from core.models import ROLE_GDTC, ROLE_HDTN, ROLE_NANG, ROLE_NANG_KEP, ROLE_THUONG
from data import repository as repo

# Standard subjects template when importing into fresh database
STANDARD_SUBJECTS = [
    ("Toán học", ROLE_NANG_KEP, 0),
    ("Ngữ văn", ROLE_NANG_KEP, 1),
    ("Ngoại ngữ", ROLE_NANG, 2),
    ("Khoa học tự nhiên (Vật lý)", ROLE_THUONG, 3),
    ("Khoa học tự nhiên (Hóa học)", ROLE_THUONG, 4),
    ("Khoa học tự nhiên (Sinh học)", ROLE_THUONG, 5),
    ("Lịch sử và Địa Lý (Lịch sử)", ROLE_THUONG, 6),
    ("Lịch sử và Địa Lý (Địa lý)", ROLE_THUONG, 7),
    ("GDCD", ROLE_THUONG, 8),
    ("Công nghệ", ROLE_THUONG, 9),
    ("Tin học", ROLE_THUONG, 10),
    ("Giáo dục thể chất", ROLE_GDTC, 11),
    ("Nội dung giáo dục của địa phương", ROLE_THUONG, 12),
    ("Hoạt động trải nghiệm, hướng nghiệp", ROLE_HDTN, 13),
    ("Nghệ thuật (Âm nhạc)", ROLE_THUONG, 14),
    ("Nghệ thuật (Mỹ thuật)", ROLE_THUONG, 15),
    ("Chào cờ", ROLE_THUONG, 16),
    ("Sinh hoạt lớp", ROLE_THUONG, 17),
]

DEFAULT_GRADE_CLASSES = {
    6: ["6A5", "6A6"],
    7: ["7A4", "7A5"],
    8: ["8A5", "8A6"],
    9: ["9A5", "9A6"],
}


def map_subject_name(mon: Optional[str], phan_mon: Optional[str] = None) -> Optional[str]:
    m = str(mon).strip() if mon is not None else ""
    pm = str(phan_mon).strip() if phan_mon is not None else ""

    # Check compound subjects
    if m in ["LS&ĐL", "Lịch sử và Địa lí", "Lịch sử và Địa lý", "LS-ĐL", "LS&DL"]:
        if "sử" in pm.lower() or "su" in pm.lower():
            return "Lịch sử và Địa Lý (Lịch sử)"
        if "địa" in pm.lower() or "dia" in pm.lower():
            return "Lịch sử và Địa Lý (Địa lý)"
    if m in ["KHTN", "Khoa học tự nhiên"]:
        if any(x in pm.lower() for x in ["vật", "vat", "lí", "li", "lý", "ly"]):
            return "Khoa học tự nhiên (Vật lý)"
        if any(x in pm.lower() for x in ["hoá", "hoa", "hóa"]):
            return "Khoa học tự nhiên (Hóa học)"
        if "sinh" in pm.lower():
            return "Khoa học tự nhiên (Sinh học)"
    if m in ["NT", "Nghệ thuật"]:
        if any(x in pm.lower() for x in ["an", "ân", "âm", "nhạc", "nhac"]):
            return "Nghệ thuật (Âm nhạc)"
        if any(x in pm.lower() for x in ["mt", "mỹ", "mĩ", "thuật", "thuat"]):
            return "Nghệ thuật (Mỹ thuật)"
    if any(x in m.lower() for x in ["địa phương", "dia phuong", "gdđp", "gddp"]):
        return "Nội dung giáo dục của địa phương"
    if any(x in m.lower() for x in ["hđtn", "hdtn", "hđ hn-tn", "hd hn-tn", "trải nghiệm", "trai nghiem"]):
        return "Hoạt động trải nghiệm, hướng nghiệp"
    if any(x in m.lower() for x in ["gdtc", "thể chất", "the chat"]):
        return "Giáo dục thể chất"
    if any(x in m.lower() for x in ["gdcd", "công dân", "cong dan"]):
        return "GDCD"
    if "tin" in m.lower():
        return "Tin học"
    if "công nghệ" in m.lower() or "cong nghe" in m.lower():
        return "Công nghệ"
    if "toán" in m.lower() or "toan" in m.lower():
        return "Toán học"
    if any(x in m.lower() for x in ["văn", "van", "ngữ văn", "ngu van"]):
        return "Ngữ văn"
    if any(x in m.lower() for x in ["anh", "ngoại ngữ", "ngoai ngu", "tiếng anh"]):
        return "Ngoại ngữ"
    return None


def _find_grade_from_sheet_name(sheet_name: str) -> Optional[int]:
    name = sheet_name.strip().upper()
    # Skip summary sheets
    if "KHTN" in name or "LS&ĐL" in name:
        return None
    m = re.search(r"K([6-9])", name)
    if m:
        return int(m.group(1))
    m2 = re.search(r"KHỐI\s*([6-9])", name, re.IGNORECASE)
    if m2:
        return int(m2.group(1))
    return None


def import_weekly_curriculum_from_excel(
    conn: sqlite3.Connection,
    file_source: Union[str, bytes, BinaryIO, openpyxl.Workbook],
) -> dict:
    """Parses Excel workbook and stores full-year weekly periods for all classes."""
    if isinstance(file_source, openpyxl.Workbook):
        wb = file_source
    elif isinstance(file_source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(file_source), data_only=True)
    elif hasattr(file_source, "read"):
        wb = openpyxl.load_workbook(file_source, data_only=True)
    else:
        wb = openpyxl.load_workbook(str(file_source), data_only=True)

    # Ensure standard subjects exist if subjects table is empty
    db_subjects = repo.list_subjects(conn)
    if not db_subjects:
        for sname, role_code, sorder in STANDARD_SUBJECTS:
            repo.upsert_subject(conn, sname, role_code=role_code, sort_order=sorder)
        db_subjects = repo.list_subjects(conn)

    subj_name_to_id = {s.name: s.subject_id for s in db_subjects}

    # Ensure standard classes exist if classes table is empty or missing grades
    db_classes = repo.list_classes(conn)
    grade_to_classes: dict[int, list] = {6: [], 7: [], 8: [], 9: []}
    for c in db_classes:
        m = re.match(r"^(\d+)", c.name)
        if m:
            g = int(m.group(1))
            if g in grade_to_classes:
                grade_to_classes[g].append(c)

    # Auto-provision classes for any empty grade
    sort_idx = len(db_classes)
    for g, def_names in DEFAULT_GRADE_CLASSES.items():
        if not grade_to_classes[g]:
            for cname in def_names:
                repo.upsert_class(conn, cname, sort_order=sort_idx)
                sort_idx += 1
    
    db_classes = repo.list_classes(conn)
    grade_to_classes = {6: [], 7: [], 8: [], 9: []}
    for c in db_classes:
        m = re.match(r"^(\d+)", c.name)
        if m:
            g = int(m.group(1))
            if g in grade_to_classes:
                grade_to_classes[g].append(c)

    all_parsed_entries = []
    classes_updated = set()
    subjects_mapped = set()
    all_weeks_found = set()
    grades_found = set()

    for sheet_name in wb.sheetnames:
        grade = _find_grade_from_sheet_name(sheet_name)
        if grade is None or grade not in grade_to_classes or not grade_to_classes[grade]:
            continue

        grades_found.add(grade)
        ws = wb[sheet_name]

        # 1. Find week header row
        week_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and 1 <= int(v) <= 52:
                    v_next = ws.cell(r, c + 1).value
                    if isinstance(v_next, (int, float)) and int(v_next) == int(v) + 1:
                        week_row = r
                        break
            if week_row is not None:
                break

        if week_row is None:
            continue

        # Build column index to week number mapping
        col_to_week = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(week_row, c).value
            if isinstance(v, (int, float)) and 1 <= int(v) <= 52:
                w_int = int(v)
                col_to_week[c] = w_int
                all_weeks_found.add(w_int)

        if not col_to_week:
            continue

        # 2. Iterate subject rows
        last_mon = None
        for r in range(week_row + 1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = ws.cell(r, 2).value
            c3 = ws.cell(r, 3).value

            # Stop at total rows
            if (
                (c1 and str(c1).strip().lower().startswith("tổng"))
                or (c2 and str(c2).strip().lower().startswith("tổng"))
            ):
                break

            if c2 is not None and str(c2).strip():
                last_mon = str(c2).strip()

            canonical_name = map_subject_name(last_mon, c3)
            if not canonical_name:
                continue

            if canonical_name not in subj_name_to_id:
                # Dynamically create subject if missing -- ROLE_THUONG (môn thường, không
                # heavy/kép/GDTC/HDTN) là mặc định an toàn cho môn mới phát hiện qua import,
                # trường có thể sửa lại role_code thủ công sau ở trang Khai báo nếu cần
                # (2026-09-05: sửa NameError -- ROLE_NONE chưa từng tồn tại/import ở đây).
                sid = repo.upsert_subject(conn, canonical_name, role_code=ROLE_THUONG, sort_order=len(subj_name_to_id))
                subj_name_to_id[canonical_name] = sid

            subj_id = subj_name_to_id[canonical_name]
            subjects_mapped.add(canonical_name)

            for col_idx, w_num in col_to_week.items():
                raw_val = ws.cell(r, col_idx).value
                try:
                    val = int(float(raw_val)) if raw_val is not None and str(raw_val).strip() != "" else 0
                except (ValueError, TypeError):
                    val = 0

                for cls in grade_to_classes[grade]:
                    all_parsed_entries.append((subj_id, cls.class_id, w_num, val))
                    classes_updated.add(cls.name)

    if all_parsed_entries:
        repo.bulk_set_weekly_curriculum(conn, all_parsed_entries)

        # Synchronize periods_per_week (Chẵn/Lẻ)
        # We use Week 1 for Odd ('L') and Week 2 for Even ('C') representative of Semester 1
        # so that subjects like Công nghệ (Khối 8, 9) accurately reflect 2 periods in HKI!
        parsed_dict = {(s_id, c_id, w): p for s_id, c_id, w, p in all_parsed_entries}
        for cls in db_classes:
            for subj in db_subjects:
                # Find odd week period (prefer week 1, 3, 5...)
                val_l = None
                for w in [1, 3, 5, 7, 9]:
                    if (subj.subject_id, cls.class_id, w) in parsed_dict:
                        val_l = parsed_dict[(subj.subject_id, cls.class_id, w)]
                        break
                if val_l is None:
                    odd_vals = [p for (s, c, w), p in parsed_dict.items() if s == subj.subject_id and c == cls.class_id and w % 2 != 0]
                    if odd_vals:
                        val_l = odd_vals[0]

                # Find even week period (prefer week 2, 4, 6, 8...)
                val_c = None
                for w in [2, 4, 6, 8]:
                    if (subj.subject_id, cls.class_id, w) in parsed_dict:
                        val_c = parsed_dict[(subj.subject_id, cls.class_id, w)]
                        break
                if val_c is None:
                    even_vals = [p for (s, c, w), p in parsed_dict.items() if s == subj.subject_id and c == cls.class_id and w % 2 == 0]
                    if even_vals:
                        val_c = even_vals[0]

                if val_l is not None:
                    repo.set_periods_per_week(conn, subj.subject_id, cls.class_id, "L", int(val_l))
                if val_c is not None:
                    repo.set_periods_per_week(conn, subj.subject_id, cls.class_id, "C", int(val_c))

    return {
        "records_imported": len(all_parsed_entries),
        "weeks_count": len(all_weeks_found),
        "weeks": sorted(all_weeks_found),
        "grades_found": sorted(grades_found),
        "classes_updated": sorted(classes_updated),
        "subjects_mapped": sorted(subjects_mapped),
    }
