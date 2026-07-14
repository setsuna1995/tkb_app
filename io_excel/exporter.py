"""Export the current DB state (or a specific accepted run) to a downloadable
.xlsx, using export_template.xlsm (a copy of the school's own TKB_9lop_moi.xlsm)
as the base -- so the output keeps the same fonts, header colors, borders and
alternating row banding as the original workbook. Only the TKB_Mon (subject
only, renamed from the template's TKB_Nhap) / TKB (subject + teacher) /
TKB_GV (per-teacher, conflict-highlighted) sheets are rewritten; column
widths/row heights are auto-fit to content instead of kept from the
template. The bundled template file itself is loaded fresh every call and
never mutated. VBA macros are dropped on load (superseded by this app),
keeping the output a plain .xlsx.
"""
from __future__ import annotations

import io
import os
from collections import defaultdict
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

from core import frame as frame_mod
from core.models import WEEKDAY_NAMES, WEEKDAYS
from data import repository as repo

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "export_template.xlsm")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

N_GRID_COLS = 3 + len(WEEKDAYS) + 1  # LỚP HỌC, BUỔI, TIẾT THỨ, Thứ 2..7, CHỦ NHẬT


def _capture_row_style(ws, row_idx: int, n_cols: int) -> list:
    return [
        (copy(c.font), copy(c.fill), copy(c.border), copy(c.alignment), c.number_format)
        for c in (ws.cell(row_idx, col) for col in range(1, n_cols + 1))
    ]


def _apply_row_style(ws, row_idx: int, style: list) -> None:
    for col, (font, fill, border, alignment, number_format) in enumerate(style, start=1):
        cell = ws.cell(row_idx, col)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = alignment
        cell.number_format = number_format


def _detect_banding(ws, first_data_row: int, n_cols: int, max_scan: int = 200) -> tuple:
    """(style_a, style_b): the two alternating row styles the template uses for
    class-block banding (style_a for the first class, style_b for the next...).
    Falls back to (style_a, style_a) -- no banding -- if none is found.
    """
    style_a = _capture_row_style(ws, first_data_row, n_cols)
    for r in range(first_data_row + 1, min(first_data_row + max_scan, ws.max_row) + 1):
        fill = ws.cell(r, 1).fill
        if fill and fill.fill_type == "solid" and fill.fgColor.rgb not in (None, "00000000"):
            return style_a, _capture_row_style(ws, r, n_cols)
    return style_a, style_a


def _clear_values(ws, first_data_row: int) -> None:
    for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _autofit_sheet(ws, min_width: int = 8, max_width: int = 40, col_padding: int = 2,
                    row_height_per_line: float = 15, min_row_height: float = 15) -> None:
    """Xấp xỉ auto-fit độ rộng cột + chiều cao hàng theo nội dung thực tế (openpyxl không có
    autofit thật vì không render text). Cell nhiều dòng ("môn\\nGV: tên") tính theo dòng dài
    nhất cho độ rộng cột, và theo số dòng cho chiều cao hàng.
    """
    col_widths: dict = {}
    row_lines: dict = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            lines = str(cell.value).split("\n")
            longest_line = max(len(line) for line in lines)
            col_widths[cell.column_letter] = max(col_widths.get(cell.column_letter, 0), longest_line)
            row_lines[cell.row] = max(row_lines.get(cell.row, 1), len(lines))
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(width + col_padding, max_width))
    for row_idx, n_lines in row_lines.items():
        ws.row_dimensions[row_idx].height = max(min_row_height, n_lines * row_height_per_line)


def _fill_result_sheets(ws_raw, ws_tkb, ws_gv, cells, classes, frame_templates, assignments,
                         teacher_names, subject_names) -> None:
    """Điền dữ liệu 1 tuần (1 parity) vào bộ 3 sheet TKB_Mon (chỉ tên môn)/TKB (môn+GV)/TKB_GV
    (theo GV, tô đỏ trùng lịch) đã cho."""
    # find teacher double-bookings (same teacher, same weekday+session+period, across classes)
    slot_teacher_classes = defaultdict(list)
    for (cid, wd, sess, per), subj_id in cells.items():
        if subj_id is None:
            continue
        teacher_id = assignments.get((subj_id, cid))
        if teacher_id is not None:
            slot_teacher_classes[(teacher_id, wd, sess, per)].append(cid)
    conflicts = {key for key, cls_list in slot_teacher_classes.items() if len(cls_list) > 1}

    for ws in (ws_raw, ws_tkb, ws_gv):
        white_style, gray_style = _detect_banding(ws, first_data_row=2, n_cols=N_GRID_COLS)
        _clear_values(ws, first_data_row=2)

        row_idx = 2
        for class_idx, cls in enumerate(classes):
            # morning/afternoon ở đây là số tiết CHUẨN (chưa trừ ngày lệch) -- dùng để sinh đủ
            # hàng cho mọi tiết có thể xuất hiện trong tuần, kể cả những ngày không phải ngày lệch.
            morning, afternoon, _study_sunday, _allow_saturday, _short_wd, _short_m, _short_a = \
                frame_templates.get(cls.class_id, (5, 3, 0, 0, None, None, None))
            sessions = [("S", p) for p in range(1, morning + 1)] + [("C", p) for p in range(1, afternoon + 1)]
            band_style = white_style if class_idx % 2 == 0 else gray_style
            for session, period in sessions:
                _apply_row_style(ws, row_idx, band_style)
                ws.cell(row_idx, 1).value = cls.name
                ws.cell(row_idx, 2).value = session
                ws.cell(row_idx, 3).value = period
                for i, wd in enumerate(WEEKDAYS):
                    subj_id = cells.get((cls.class_id, wd, session, period))
                    subj_name = subject_names.get(subj_id, "") if subj_id else ""
                    teacher_id = assignments.get((subj_id, cls.class_id)) if subj_id else None
                    teacher_name = teacher_names.get(teacher_id, "") if teacher_id else ""
                    col = 4 + i
                    if ws is ws_gv:
                        ws.cell(row_idx, col).value = teacher_name
                        if teacher_id is not None and (teacher_id, wd, session, period) in conflicts:
                            ws.cell(row_idx, col).fill = RED_FILL
                    elif ws is ws_tkb:
                        ws.cell(row_idx, col).value = (
                            f"{subj_name}\nGV: {teacher_name or '(chưa PC GV)'}" if subj_name else ""
                        )
                    else:  # ws_raw
                        ws.cell(row_idx, col).value = subj_name
                row_idx += 1


def export_xlsx(conn, run_id=None) -> bytes:
    classes = repo.list_classes(conn)
    subjects = repo.list_subjects(conn)
    teacher_names = {t.teacher_id: t.name for t in repo.list_teachers(conn)}
    subject_names = {s.subject_id: s.name for s in subjects}
    assignments = repo.get_assignments(conn)
    frame_templates = repo.get_all_frame_templates(conn)

    if run_id is not None:
        cells = repo.get_tkb_result(conn, run_id)
    else:
        cells = repo.get_tkb_nhap(conn)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)  # drop VBA -- superseded by this app
    ws_raw = wb["TKB_Nhap"]
    ws_raw.title = "TKB_Mon"
    # sheet môn-only vốn có dropdown chọn môn (cột D:J) trỏ tới defined-name DS_Mon =
    # PhanCong!$A$3:$A$18 -- PhanCong đã bị xoá khỏi file xuất (stale, không refresh từ DB)
    # nên link này hỏng (#REF!/cảnh báo "broken link" khi mở Excel thật). Gỡ bỏ hẳn.
    ws_raw.data_validations.dataValidation.clear()
    if "DS_Mon" in wb.defined_names:
        del wb.defined_names["DS_Mon"]
    ws_tkb = wb["TKB"]
    ws_gv = wb["TKB_GV"]

    # the template also carries PhanCong/SoTiet/DinhMuc_GV/KiemTra/... -- những sheet đó sẽ
    # stale (không refresh từ DB) hoặc không còn dùng, nên xoá hết, chỉ giữ 3 sheet kết quả.
    keep = {"TKB_Mon", "TKB", "TKB_GV"}
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]

    _fill_result_sheets(ws_raw, ws_tkb, ws_gv, cells, classes, frame_templates, assignments,
                         teacher_names, subject_names)
    for ws in (ws_raw, ws_tkb, ws_gv):
        _autofit_sheet(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


PARITY_SUFFIX = {"C": "Chan", "L": "Le"}
PARITY_LABEL = {"C": "Chẵn", "L": "Lẻ"}


def export_xlsx_both_parities(conn) -> tuple:
    """Gộp lần chấp nhận gần nhất của MỖI tuần (Chẵn + Lẻ) vào 1 workbook 6 sheet
    (3 sheet hiện có, hậu tố _Chan/_Le). tkb_nhap không dùng được ở đây vì nó chỉ lưu
    1 bản duy nhất, bị ghi đè mỗi lần chấp nhận bất kể tuần nào -- nguồn dữ liệu đúng
    cho từng tuần là run_log/tkb_result (không bao giờ bị xoá, có cột parity).

    Trả về (bytes, warnings) -- warnings liệt kê tuần nào bị bỏ qua vì chưa từng có
    lần xếp nào được chấp nhận. Raise ValueError nếu CẢ 2 tuần đều chưa có gì để xuất.
    """
    classes = repo.list_classes(conn)
    subjects = repo.list_subjects(conn)
    teacher_names = {t.teacher_id: t.name for t in repo.list_teachers(conn)}
    subject_names = {s.subject_id: s.name for s in subjects}
    assignments = repo.get_assignments(conn)
    frame_templates = repo.get_all_frame_templates(conn)

    warnings = []
    parity_cells = {}
    for parity in ("C", "L"):
        run = repo.get_latest_run_by_parity(conn, parity)
        if run is None:
            warnings.append(f"Tuần {PARITY_LABEL[parity]}: chưa có lần xếp nào được chấp nhận -- bỏ qua.")
            continue
        parity_cells[parity] = repo.get_tkb_result(conn, run["run_id"])

    if not parity_cells:
        raise ValueError(
            "Chưa có lần xếp nào được chấp nhận cho tuần Chẵn hoặc tuần Lẻ -- không có gì để xuất."
        )

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    base_sheet_map = {"TKB_Mon": "TKB_Nhap", "TKB": "TKB", "TKB_GV": "TKB_GV"}
    base_sheets = {out_name: wb[tmpl_name] for out_name, tmpl_name in base_sheet_map.items()}
    for name in list(wb.sheetnames):
        if name not in base_sheet_map.values():
            del wb[name]

    # Gỡ link hỏng trên sheet gốc TRƯỚC khi copy, để cả 2 bản copy (Chẵn + Lẻ) đều sạch --
    # xem chú thích tương tự trong export_xlsx().
    base_sheets["TKB_Mon"].data_validations.dataValidation.clear()
    if "DS_Mon" in wb.defined_names:
        del wb.defined_names["DS_Mon"]

    for parity, cells in parity_cells.items():
        suffix = PARITY_SUFFIX[parity]
        copies = {}
        for out_name, base_ws in base_sheets.items():
            new_ws = wb.copy_worksheet(base_ws)
            new_ws.title = f"{out_name}_{suffix}"
            new_ws.freeze_panes = base_ws.freeze_panes  # copy_worksheet không giữ freeze_panes
            copies[out_name] = new_ws
        _fill_result_sheets(copies["TKB_Mon"], copies["TKB"], copies["TKB_GV"], cells, classes,
                             frame_templates, assignments, teacher_names, subject_names)
        for ws in copies.values():
            _autofit_sheet(ws)

    for base_ws in base_sheets.values():
        wb.remove(base_ws)  # chỉ dùng làm khuôn để copy_worksheet, không cần trong output; xoá
        # theo object (không theo tên) vì tên output ("TKB_Mon") đã lệch tên sheet gốc ("TKB_Nhap")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), warnings


def export_full_backup_xlsx(conn) -> bytes:
    """Xuất TOÀN BỘ dữ liệu setup (không chỉ lịch) ra 1 file .xlsm mà import_xlsm() đọc lại
    được nguyên vẹn. Khác với export_xlsx()/export_xlsx_both_parities() (chỉ xuất lưới thời
    khóa biểu để in/chia sẻ) -- hàm này dùng cho nút "sao lưu", vì mất DB (vd host free-tier
    restart) mà chỉ có lưới TKB thì không khôi phục lại được lớp/môn/GV/phân công/định mức/
    GV bận/khung tiết/lịch sử tuần, phải nhập tay lại từ đầu.

    Ghi đúng 7 sheet, đúng định dạng io_excel/importer.py::import_xlsm() mong đợi: PhanCong,
    SoTiet, DinhMuc_GV, GV_Ban, Khung, TKB_Nhap, TuanConfig.
    """
    classes = sorted(repo.list_classes(conn), key=lambda c: c.sort_order)
    subjects = sorted(repo.list_subjects(conn), key=lambda s: s.sort_order)
    teachers = repo.list_teachers(conn)
    teacher_names = {t.teacher_id: t.name for t in teachers}
    subject_names = {s.subject_id: s.name for s in subjects}

    assignments = repo.get_assignments(conn)
    periods_per_week = repo.get_periods_per_week(conn)
    role_reduction = repo.get_role_reduction(conn)
    frame_templates = repo.get_all_frame_templates(conn)
    unavailability = repo.list_unavailability(conn)
    tkb_nhap = repo.get_tkb_nhap(conn)
    seed, parity = repo.get_tuan_config(conn)
    seed_history = repo.list_seed_history(conn)
    base_cap = repo.get_base_cap(conn)
    min_floor = repo.get_min_floor(conn)

    n_classes = len(classes)
    n_subjects = len(subjects)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    keep = {"PhanCong", "SoTiet", "DinhMuc_GV", "GV_Ban", "Khung", "TKB_Nhap", "TuanConfig"}
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]

    # ---- PhanCong ----
    ws_pc = wb["PhanCong"]
    _clear_values(ws_pc, first_data_row=2)
    code_col = 2 + n_classes + 1  # để trống 1 cột đệm ở 2+n_classes, khớp quy ước template
    for i, cls in enumerate(classes):
        ws_pc.cell(2, 2 + i).value = cls.name
    ws_pc.cell(2, code_col).value = "MÃ VAI TRÒ"
    for r, subj in enumerate(subjects):
        row = 3 + r
        ws_pc.cell(row, 1).value = subj.name
        ws_pc.cell(row, code_col).value = subj.role_code
        for i, cls in enumerate(classes):
            teacher_id = assignments.get((subj.subject_id, cls.class_id))
            ws_pc.cell(row, 2 + i).value = teacher_names.get(teacher_id, "")

    # ---- SoTiet ----
    ws_st = wb["SoTiet"]
    _clear_values(ws_st, first_data_row=2)
    odd_start_col = 2 + n_classes + 1
    for i, cls in enumerate(classes):
        ws_st.cell(2, 2 + i).value = f"{cls.name} C"
        ws_st.cell(2, odd_start_col + i).value = f"{cls.name} L"
    for r, subj in enumerate(subjects):
        row = 3 + r
        ws_st.cell(row, 1).value = subj.name
        for i, cls in enumerate(classes):
            ws_st.cell(row, 2 + i).value = periods_per_week.get((subj.subject_id, cls.class_id, "C"), 0)
            ws_st.cell(row, odd_start_col + i).value = periods_per_week.get((subj.subject_id, cls.class_id, "L"), 0)

    # ---- DinhMuc_GV ----
    ws_dm = wb["DinhMuc_GV"]
    _clear_values(ws_dm, first_data_row=2)  # row1 (tiêu đề + Chuẩn/Chức vụ) giữ nguyên, ghi đè bên dưới
    ws_dm.cell(1, 8).value = "Chuẩn:"
    ws_dm.cell(1, 9).value = base_cap
    ws_dm.cell(1, 11).value = "Chức vụ"
    ws_dm.cell(1, 12).value = "Giảm"
    # Sàn tối thiểu: quy ước RIÊNG của app (không có trong định dạng VBA gốc) -- đặt ở ô N1/O1,
    # rõ ràng chưa dùng cho mục đích nào khác trong sheet này.
    ws_dm.cell(1, 14).value = "Sàn tối thiểu:"
    ws_dm.cell(1, 15).value = min_floor
    ws_dm.cell(2, 1).value = "Tên GV"
    ws_dm.cell(2, 2).value = "Chức vụ"
    ws_dm.cell(2, 8).value = "Đi T2 (1/0)"
    ws_dm.cell(2, 9).value = "GVCN (1/0)"
    for r, t in enumerate(teachers):
        row = 3 + r
        ws_dm.cell(row, 1).value = t.name
        ws_dm.cell(row, 2).value = t.role
        # cột C-G (Giảm/Trần/Tải Chẵn/Tải Lẻ/Vượt) là công thức Excel, import_xlsm() không đọc
        # lại -- để trống thay vì cố tính lại, tránh hiện số liệu sai/cũ.
        ws_dm.cell(row, 8).value = int(t.must_monday)
        ws_dm.cell(row, 9).value = int(t.is_gvcn)
    for r, (role_name, reduction) in enumerate(role_reduction.items()):
        row = 2 + r
        ws_dm.cell(row, 11).value = role_name
        ws_dm.cell(row, 12).value = reduction

    # ---- GV_Ban ----
    ws_gb = wb["GV_Ban"]
    _clear_values(ws_gb, first_data_row=2)
    ws_gb.cell(2, 1).value = "Giáo viên"
    ws_gb.cell(2, 2).value = "Thứ"
    ws_gb.cell(2, 3).value = "Buổi"
    ws_gb.cell(2, 4).value = "Tiết"
    for r, row_data in enumerate(unavailability):
        row = 3 + r
        ws_gb.cell(row, 1).value = teacher_names.get(row_data["teacher_id"], "")
        ws_gb.cell(row, 2).value = row_data["weekday"]
        ws_gb.cell(row, 3).value = row_data["session"]
        ws_gb.cell(row, 4).value = row_data["period"]

    # ---- Khung + TKB_Nhap (row-aligned, importer đọc Khung theo đúng row của TKB_Nhap) ----
    ws_khung = wb["Khung"]
    ws_nh = wb["TKB_Nhap"]
    _clear_values(ws_khung, first_data_row=2)
    _clear_values(ws_nh, first_data_row=2)
    ws_nh.cell(1, 1).value = "LỚP HỌC"
    ws_nh.cell(1, 2).value = "BUỔI"
    ws_nh.cell(1, 3).value = "TIẾT THỨ"
    for i, wd in enumerate(WEEKDAYS):
        ws_nh.cell(1, 4 + i).value = WEEKDAY_NAMES[wd]
    ws_nh.cell(1, 4 + len(WEEKDAYS)).value = WEEKDAY_NAMES[8]

    row_idx = 2
    for cls in classes:
        morning, afternoon, study_sunday, allow_saturday, short_wd, short_m, short_a = \
            frame_templates.get(cls.class_id, (5, 3, 0, 0, None, None, None))
        # active_cells() đã tự xử lý đúng ngày lệch tiết -- tái dùng thẳng, không tự suy luận lại.
        active_set = set(frame_mod.active_cells(
            morning, afternoon, bool(study_sunday), bool(allow_saturday), short_wd, short_m, short_a,
        ))
        sessions = [("S", p) for p in range(1, morning + 1)] + [("C", p) for p in range(1, afternoon + 1)]
        for session, period in sessions:
            ws_nh.cell(row_idx, 1).value = cls.name
            ws_nh.cell(row_idx, 2).value = session
            ws_nh.cell(row_idx, 3).value = period
            for i, wd in enumerate(WEEKDAYS):
                col = 4 + i
                subj_id = tkb_nhap.get((cls.class_id, wd, session, period))
                ws_nh.cell(row_idx, col).value = subject_names.get(subj_id, "") if subj_id else ""
                if (wd, session, period) in active_set:
                    ws_khung.cell(row_idx, col).value = "x"
            row_idx += 1

    # Sửa lại defined-name DS_Mon (dropdown chọn môn trên TKB_Nhap) theo đúng số môn thực tế --
    # range cũ trong template (PhanCong!$A$3:$A$18) giả định đúng 16 môn mẫu của trường mẫu.
    if "DS_Mon" in wb.defined_names:
        wb.defined_names["DS_Mon"].value = f"PhanCong!$A$3:$A${2 + n_subjects}"

    # ---- TuanConfig ----
    ws_tc = wb["TuanConfig"]
    _clear_values(ws_tc, first_data_row=4)
    ws_tc.cell(1, 2).value = seed
    ws_tc.cell(2, 2).value = parity
    for r, h in enumerate(seed_history):
        row = 4 + r
        ws_tc.cell(row, 1).value = h["week_no"]
        ws_tc.cell(row, 2).value = h["seed"]
        # _extract_parity() bên importer chỉ dò chuỗi con "[C]"/"[L]" trong text này, không có
        # cột parity riêng -- phải tự chèn tag khi ghi để đọc lại đúng.
        ws_tc.cell(row, 3).value = f"{h['created_at']} [{h['parity']}]"

    for ws in (ws_pc, ws_st, ws_dm, ws_gb, ws_khung, ws_nh, ws_tc):
        _autofit_sheet(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
