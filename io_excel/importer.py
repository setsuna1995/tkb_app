"""Import an existing TKB .xlsm workbook into the SQLite schema.

Mirrors the VBA's own dynamic header-scanning (never hardcodes row/col
numbers for dimensions) so it stays robust if the school adds a class or
subject later. Skips columns that are Excel-formula-derived in the original
workbook (DinhMuc_GV's Trần/Tải/Vượt) since those are recomputed on read.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

from core.models import ROLE_HDTN
from data import repository as repo


@dataclass
class ImportReport:
    counts: dict = field(default_factory=dict)
    warnings: list = field(default_factory=list)


def _norm(value) -> str:
    return str(value).strip() if value is not None else ""


def _extract_parity(text: str) -> str:
    return "L" if "[L]" in text.upper() else "C"


def _infer_session_frame(per_weekday: dict) -> tuple:
    """per_weekday: {weekday: max_active_period} cho 1 (lớp, buổi), đọc từ lưới "x" của
    sheet Khung. Trả về (standard_periods, short_weekday, short_periods) -- 2 mục sau là
    None nếu mọi ngày đồng nhất, hoặc nếu có nhiều hơn 1 ngày lệch chuẩn (vượt quá mô hình
    1-ngày-lệch, gọi nơi dùng nên fallback về max() như hành vi cũ).
    """
    values = list(per_weekday.values())
    if not values or max(values) == 0:
        return 0, None, None
    standard = max(set(values), key=values.count)
    outliers = {wd: v for wd, v in per_weekday.items() if v != standard}
    if not outliers:
        return standard, None, None
    if len(outliers) == 1:
        (wd, v), = outliers.items()
        return standard, wd, v
    return max(values), None, None


def _resolve_frame(per_weekday_s: dict, per_weekday_c: dict) -> tuple:
    """Gộp kết quả suy luận của 2 buổi thành 1 frame_template.
    Trả về (morning_periods, afternoon_periods, short_weekday, short_morning, short_afternoon).
    """
    std_s, out_wd_s, out_val_s = _infer_session_frame(per_weekday_s)
    std_c, out_wd_c, out_val_c = _infer_session_frame(per_weekday_c)
    if out_wd_s is not None and out_wd_c is not None and out_wd_s != out_wd_c:
        # Ngày lệch của buổi sáng và buổi chiều khác nhau -- vượt quá mô hình 1 ngày lệch
        # chung cho cả lớp, fallback về khung đồng nhất (dùng periods tối đa từng buổi).
        fallback_s = max(per_weekday_s.values()) if per_weekday_s else 0
        fallback_c = max(per_weekday_c.values()) if per_weekday_c else 0
        return fallback_s, fallback_c, None, None, None
    short_wd = out_wd_s if out_wd_s is not None else out_wd_c
    short_m = out_val_s if out_wd_s is not None else None
    short_a = out_val_c if out_wd_c is not None else None
    return std_s, std_c, short_wd, short_m, short_a


def import_xlsm(conn, path: str) -> ImportReport:
    wb = openpyxl.load_workbook(path, data_only=True)
    report = ImportReport()

    ws_pc = wb["PhanCong"]
    ws_st = wb["SoTiet"]

    # ---- dims: classes (row 2 from col B), subjects (col A from row 3) ----
    class_names = []
    col = 2
    while True:
        v = _norm(ws_pc.cell(2, col).value)
        if not v or v.upper().startswith("MA"):
            break
        class_names.append(v)
        col += 1
    n_classes = len(class_names)

    subject_rows = []
    row = 3
    while _norm(ws_pc.cell(row, 1).value):
        subject_rows.append((row, _norm(ws_pc.cell(row, 1).value)))
        row += 1
    n_subjects = len(subject_rows)

    if n_classes == 0 or n_subjects == 0:
        raise ValueError(
            "PhanCong rỗng: không đọc được lớp (hàng 2 từ cột B) hoặc môn (cột A từ hàng 3)."
        )

    # ---- role-code ("MA") column: dynamic scan, same as ResolveRoles ----
    code_col = None
    for c in range(2 + n_classes, 2 + n_classes + 6):
        if _norm(ws_pc.cell(2, c).value).upper().startswith("MA"):
            code_col = c
            break
    if code_col is None:
        code_col = 2 + n_classes + 1

    class_ids = {}
    for i, name in enumerate(class_names):
        existing_class_id = repo.get_class_by_name(conn, name)
        class_ids[name] = repo.upsert_class(conn, name, sort_order=i, class_id=existing_class_id)

    subject_ids = {}
    hdtn_present = False
    for i, (row_idx, name) in enumerate(subject_rows):
        role_code = int(ws_pc.cell(row_idx, code_col).value or 0)
        existing_subject_id = repo.get_subject_by_name(conn, name)
        subject_ids[name] = repo.upsert_subject(
            conn, name, role_code=role_code, sort_order=i, subject_id=existing_subject_id
        )
        if role_code == ROLE_HDTN:
            hdtn_present = True
    if not hdtn_present:
        report.warnings.append(
            "Không tìm thấy môn có MÃ = 5 (HDTN) ở cột 'MÃ VAI TRÒ' trên PhanCong. "
            "Xếp TKB sẽ báo lỗi cho tới khi được bổ sung."
        )

    # ---- teacher assignments (PhanCong grid) ----
    teacher_ids = {}

    def get_or_create_teacher(name: str) -> int:
        if name not in teacher_ids:
            existing = repo.get_teacher_by_name(conn, name)
            teacher_ids[name] = existing if existing is not None else repo.upsert_teacher(conn, name)
        return teacher_ids[name]

    for row_idx, subj_name in subject_rows:
        for i, cls_name in enumerate(class_names):
            teacher_name = _norm(ws_pc.cell(row_idx, 2 + i).value)
            teacher_id = get_or_create_teacher(teacher_name) if teacher_name else None
            repo.set_assignment(conn, subject_ids[subj_name], class_ids[cls_name], teacher_id)

    # ---- SoTiet: even (Chẵn) block from col B, odd (Lẻ) block from col (2+n_classes+1) ----
    odd_start_col = 2 + n_classes + 1
    for row_idx, subj_name in subject_rows:
        for i, cls_name in enumerate(class_names):
            even_val = int(ws_st.cell(row_idx, 2 + i).value or 0)
            odd_val = int(ws_st.cell(row_idx, odd_start_col + i).value or 0)
            repo.set_periods_per_week(conn, subject_ids[subj_name], class_ids[cls_name], "C", even_val)
            repo.set_periods_per_week(conn, subject_ids[subj_name], class_ids[cls_name], "L", odd_val)

    # ---- DinhMuc_GV: only the hand-entered columns (role, Đi T2, GVCN) ----
    n_teachers_from_dm = 0
    if "DinhMuc_GV" in wb.sheetnames:
        ws_dm = wb["DinhMuc_GV"]
        row = 3
        while _norm(ws_dm.cell(row, 1).value):
            name = _norm(ws_dm.cell(row, 1).value)
            role = _norm(ws_dm.cell(row, 2).value)
            must_monday = bool(int(ws_dm.cell(row, 8).value or 0))
            is_gvcn = bool(int(ws_dm.cell(row, 9).value or 0))
            tid = get_or_create_teacher(name)
            repo.upsert_teacher(conn, name, role=role, must_monday=must_monday, is_gvcn=is_gvcn, teacher_id=tid)
            n_teachers_from_dm += 1
            row += 1

        # role -> reduction lookup table, headed by "Chức vụ"/"Giảm" (searched dynamically, not fixed columns)
        role_col_idx = None
        for c in range(1, 25):
            if _norm(ws_dm.cell(1, c).value) == "Chức vụ":
                role_col_idx = c
                break
        if role_col_idx is not None:
            r = 2
            while _norm(ws_dm.cell(r, role_col_idx).value):
                r_name = _norm(ws_dm.cell(r, role_col_idx).value)
                r_reduction = int(ws_dm.cell(r, role_col_idx + 1).value or 0)
                repo.set_role_reduction(conn, r_name, r_reduction)
                r += 1

    # ---- GV_Ban: skip rows whose name doesn't match a known teacher (also filters instructional rows) ----
    n_unavailability = 0
    if "GV_Ban" in wb.sheetnames:
        ws_gb = wb["GV_Ban"]
        row = 3
        while _norm(ws_gb.cell(row, 1).value):
            name = _norm(ws_gb.cell(row, 1).value)
            tid = repo.get_teacher_by_name(conn, name)
            if tid is None:
                report.warnings.append(f"GV_Bận dòng {row}: bỏ qua vì không khớp tên GV nào ('{name}').")
            else:
                weekday = _norm(ws_gb.cell(row, 2).value).upper() or "*"
                session = _norm(ws_gb.cell(row, 3).value).upper() or "*"
                period = _norm(ws_gb.cell(row, 4).value) or "*"
                repo.add_unavailability(conn, tid, weekday, session, period)
                n_unavailability += 1
            row += 1

    # ---- TKB_Nhap (baseline grid) + infer frame_template from the real Khung "x" pattern ----
    ws_nh = wb["TKB_Nhap"]
    ws_khung = wb["Khung"] if "Khung" in wb.sheetnames else None
    # Theo dõi period active lớn nhất TỪNG NGÀY (không gộp chung) để phát hiện đúng ngày lệch
    # tiết (vd Thứ 7 chỉ 4 tiết trong khi các ngày khác 5 tiết) thay vì lấy max() rồi áp đồng
    # nhất cho mọi ngày như trước, làm mất thông tin ngày lệch tiết của trường thực tế.
    frame_per_weekday = {
        cid: {"S": {wd: 0 for wd in range(2, 8)}, "C": {wd: 0 for wd in range(2, 8)}}
        for cid in class_ids.values()
    }
    cells = {}
    row = 2
    while _norm(ws_nh.cell(row, 1).value):
        cls_name = _norm(ws_nh.cell(row, 1).value)
        if cls_name not in class_ids:
            row += 1
            continue
        cls_id = class_ids[cls_name]
        session = _norm(ws_nh.cell(row, 2).value).upper()
        period = int(ws_nh.cell(row, 3).value or 0)

        if ws_khung is not None and session in ("S", "C"):
            for wd in range(2, 8):
                if _norm(ws_khung.cell(row, wd + 2).value):
                    per_wd = frame_per_weekday[cls_id][session]
                    per_wd[wd] = max(per_wd[wd], period)

        for wd in range(2, 8):
            val = _norm(ws_nh.cell(row, wd + 2).value)
            subj_id = subject_ids.get(val) if val else None
            cells[(cls_id, wd, session, period)] = subj_id
        row += 1

    repo.bulk_replace_tkb_nhap(conn, cells)
    class_names_by_id = {cid: name for name, cid in class_ids.items()}
    for cls_id, per_session in frame_per_weekday.items():
        morning, afternoon, short_wd, short_m, short_a = _resolve_frame(per_session["S"], per_session["C"])
        try:
            repo.set_frame_template(
                conn, cls_id, morning, afternoon, study_sunday=False,
                short_weekday=short_wd, short_morning_periods=short_m, short_afternoon_periods=short_a,
            )
        except ValueError:
            # Ngày lệch suy ra từ Khung sheet vi phạm luật "không lỗ 1 tiết" -- fallback về
            # khung đồng nhất thay vì làm hỏng cả lượt import vì 1 lớp.
            report.warnings.append(
                f"Lớp '{class_names_by_id.get(cls_id, cls_id)}': ngày lệch tiết suy ra từ sheet Khung "
                f"không hợp lệ (đúng 1 tiết lẻ) -- đã bỏ qua, dùng khung đồng nhất."
            )
            repo.set_frame_template(conn, cls_id, morning, afternoon, study_sunday=False)

    # ---- TuanConfig: current seed/parity + history ----
    n_seed_history = 0
    if "TuanConfig" in wb.sheetnames:
        ws_tc = wb["TuanConfig"]
        seed = int(ws_tc.cell(1, 2).value or 0)
        parity = _norm(ws_tc.cell(2, 2).value).upper() or "C"
        repo.set_tuan_config(conn, seed, parity)
        row = 4
        while _norm(ws_tc.cell(row, 1).value):
            week_no = int(ws_tc.cell(row, 1).value or 0)
            wk_seed = int(ws_tc.cell(row, 2).value or 0)
            created = _norm(ws_tc.cell(row, 3).value)
            repo.add_seed_history(conn, week_no, wk_seed, _extract_parity(created))
            n_seed_history += 1
            row += 1

    report.counts = {
        "classes": n_classes,
        "subjects": n_subjects,
        "teachers": len(teacher_ids),
        "unavailability_rows": n_unavailability,
        "tkb_nhap_cells": len(cells),
        "seed_history_rows": n_seed_history,
    }
    return report
