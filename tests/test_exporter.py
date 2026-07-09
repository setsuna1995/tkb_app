import io
import os

import openpyxl
import pytest

from core import scheduler as sched
from data import db, repository as repo
from io_excel.exporter import export_xlsx, export_xlsx_both_parities
from io_excel.importer import import_xlsm

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "TKB_9lop_moi.xlsm")


@pytest.fixture()
def conn(tmp_path):
    connection = db.get_connection(str(tmp_path / "test.db"))
    db.init_db(connection)
    import_xlsm(connection, FIXTURE)
    yield connection
    connection.close()


def _accept_run(conn, parity: str, seed: int, week_no: int) -> int:
    inp = repo.build_scheduling_input(conn, parity=parity, seed=seed)
    result = sched.run(inp)
    assert result.success
    cells = {}
    for slot in inp.slots:
        cells[(slot.class_id, slot.ts.weekday, slot.ts.session, slot.ts.period)] = result.assignment.get(slot.slot_id)
    run_id = repo.save_run(conn, week_no=week_no, seed=seed, parity=parity, cells_changed=result.cells_changed,
                            cells_total=result.cells_total, succeeded=True, message="ok")
    repo.save_tkb_result(conn, run_id, cells)
    return run_id


def test_export_current_baseline_has_expected_sheets(conn):
    data = export_xlsx(conn)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) == {"TKB_Mon", "TKB", "TKB_GV"}


def test_export_removes_broken_phancong_link(conn):
    # sheet môn-only (TKB_Mon, đổi tên từ TKB_Nhap của template) vốn có dropdown chọn môn trỏ tới
    # defined-name DS_Mon = PhanCong!$A$3:$A$18 -- PhanCong bị xoá khỏi file xuất nên link hỏng,
    # phải được gỡ bỏ hẳn khỏi output.
    data = export_xlsx(conn)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert list(wb["TKB_Mon"].data_validations.dataValidation) == []
    assert "DS_Mon" not in wb.defined_names


def test_export_autofits_columns_and_rows(conn):
    data = export_xlsx(conn)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["TKB"]  # sheet chắc chắn có cell 2 dòng "môn\nGV: tên"
    assert any(dim.width is not None and dim.width > 8 for dim in ws.column_dimensions.values())
    assert any(dim.height is not None and dim.height > 15 for dim in ws.row_dimensions.values())


def test_export_accepted_run_no_teacher_conflict_highlight(conn):
    parity = repo.get_tuan_config(conn)[1]
    inp = repo.build_scheduling_input(conn, parity=parity, seed=123)
    result = sched.run(inp)
    assert result.success

    cells = {}
    for slot in inp.slots:
        cells[(slot.class_id, slot.ts.weekday, slot.ts.session, slot.ts.period)] = result.assignment.get(slot.slot_id)
    run_id = repo.save_run(conn, week_no=1, seed=123, parity=parity, cells_changed=result.cells_changed,
                            cells_total=result.cells_total, succeeded=True, message="ok")
    repo.save_tkb_result(conn, run_id, cells)

    data = export_xlsx(conn, run_id=run_id)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws_gv = wb["TKB_GV"]
    for row in ws_gv.iter_rows(min_row=2, min_col=4, max_col=9):
        for cell in row:
            # no conflicts on a successful run -- just never the conflict-red fill
            # (cells still carry the template's own white/gray banding fill)
            assert cell.fill.fgColor.rgb != "00FFC7CE"


def test_export_both_parities_raises_when_neither_accepted(conn):
    with pytest.raises(ValueError):
        export_xlsx_both_parities(conn)


def test_export_both_parities_warns_when_only_one_accepted(conn):
    _accept_run(conn, "C", seed=111, week_no=1)

    data, warnings = export_xlsx_both_parities(conn)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) == {"TKB_Mon_Chan", "TKB_Chan", "TKB_GV_Chan"}
    assert len(warnings) == 1
    assert "Lẻ" in warnings[0]


def test_export_both_parities_has_all_6_sheets_when_both_accepted(conn):
    _accept_run(conn, "C", seed=111, week_no=1)
    _accept_run(conn, "L", seed=222, week_no=2)

    data, warnings = export_xlsx_both_parities(conn)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) == {
        "TKB_Mon_Chan", "TKB_Chan", "TKB_GV_Chan",
        "TKB_Mon_Le", "TKB_Le", "TKB_GV_Le",
    }
    assert warnings == []


def test_export_both_parities_preserves_freeze_panes(conn):
    _accept_run(conn, "C", seed=111, week_no=1)

    data, _warnings = export_xlsx_both_parities(conn)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # copy_worksheet() của openpyxl không tự giữ freeze_panes -- regression guard
    # cho việc gán lại thủ công trong export_xlsx_both_parities.
    assert wb["TKB_Mon_Chan"].freeze_panes == "D62"
    assert wb["TKB_Chan"].freeze_panes == "D2"
    assert wb["TKB_GV_Chan"].freeze_panes == "D2"
