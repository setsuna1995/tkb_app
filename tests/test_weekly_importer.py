"""Tests for Excel weekly curriculum importer."""
import os
import sqlite3
import pytest
from data import db
from data import repository as repo

# Standard school fixture setup
@pytest.fixture
def school_conn():
    c = db.get_connection(":memory:")
    db.init_db(c)
    
    classes = ["6A5", "6A6", "7A4", "7A5", "8A5", "8A6", "9A5", "9A6"]
    for i, name in enumerate(classes):
        repo.upsert_class(c, name, i)
        
    subjects = [
        "Toán học", "Ngữ văn", "Ngoại ngữ",
        "Khoa học tự nhiên (Vật lý)", "Khoa học tự nhiên (Hóa học)", "Khoa học tự nhiên (Sinh học)",
        "Lịch sử và Địa Lý (Lịch sử)", "Lịch sử và Địa Lý (Địa lý)",
        "GDCD", "Công nghệ", "Tin học", "Giáo dục thể chất",
        "Nội dung giáo dục của địa phương", "Hoạt động trải nghiệm, hướng nghiệp",
        "Nghệ thuật (Âm nhạc)", "Nghệ thuật (Mỹ thuật)",
    ]
    for i, name in enumerate(subjects):
        repo.upsert_subject(c, name, 0, i)
    return c


def test_import_weekly_curriculum_real_excel(school_conn):
    excel_path = "Định lượng số tiết theo tuần năm học 2026_2027.xlsx"
    if not os.path.exists(excel_path):
        pytest.skip(f"Excel file {excel_path} not found")

    from io_excel.weekly_importer import import_weekly_curriculum_from_excel

    report = import_weekly_curriculum_from_excel(school_conn, excel_path)
    assert report["records_imported"] > 0
    assert report["weeks_count"] == 35
    assert len(report["classes_updated"]) == 8

    # Verify Class 6A5 has 29 periods in all 35 weeks
    c_6a5_id = repo.get_class_by_name(school_conn, "6A5")
    assert c_6a5_id is not None
    cur_6a5 = repo.get_weekly_curriculum(school_conn, class_id=c_6a5_id)
    
    for w in range(1, 36):
        total_w = sum(cur_6a5.get((s.subject_id, c_6a5_id, w), 0) for s in repo.list_subjects(school_conn))
        assert total_w == 29, f"6A5 week {w} has {total_w} periods, expected 29"

    # Verify Class 8A5 has 30 periods in weeks 1..9 & 28..35, and 29 periods in weeks 10..27
    c_8a5_id = repo.get_class_by_name(school_conn, "8A5")
    assert c_8a5_id is not None
    cur_8a5 = repo.get_weekly_curriculum(school_conn, class_id=c_8a5_id)

    for w in range(1, 10):
        total_w = sum(cur_8a5.get((s.subject_id, c_8a5_id, w), 0) for s in repo.list_subjects(school_conn))
        assert total_w == 30, f"8A5 week {w} has {total_w} periods, expected 30"

    for w in range(10, 28):
        total_w = sum(cur_8a5.get((s.subject_id, c_8a5_id, w), 0) for s in repo.list_subjects(school_conn))
        assert total_w == 29, f"8A5 week {w} has {total_w} periods, expected 29"

    for w in range(28, 36):
        total_w = sum(cur_8a5.get((s.subject_id, c_8a5_id, w), 0) for s in repo.list_subjects(school_conn))
        assert total_w == 30, f"8A5 week {w} has {total_w} periods, expected 30"

    # Verify Cong nghe has 2 periods for 8A5 in HKI weeks 1..9 and periods_per_week
    s_cn_id = repo.get_subject_by_name(school_conn, "Công nghệ")
    assert s_cn_id is not None
    for w in range(1, 10):
        assert cur_8a5.get((s_cn_id, c_8a5_id, w), 0) == 2, f"8A5 week {w} Cong nghe must be 2 periods"

    ppw = repo.get_periods_per_week(school_conn)
    assert ppw.get((s_cn_id, c_8a5_id, "C"), 0) == 2, "8A5 Cong nghe Even week must be 2"
    assert ppw.get((s_cn_id, c_8a5_id, "L"), 0) == 2, "8A5 Cong nghe Odd week must be 2"


def test_import_creates_missing_subject_without_crashing():
    """Regression (2026-09-05): import_weekly_curriculum_from_excel dùng
    ROLE_NONE (chưa từng tồn tại/được import trong file) khi cần TỰ TẠO một
    môn chưa có sẵn trong DB -- gây NameError bất cứ khi nào 1 trường thiếu 1
    môn chuẩn hoặc import theo thứ tự khác thường. Phải tự tạo môn (role
    ROLE_THUONG) thay vì crash."""
    import openpyxl
    from data import db
    from data import repository as repo
    from core.models import ROLE_THUONG
    from io_excel.weekly_importer import import_weekly_curriculum_from_excel

    conn = db.get_connection(":memory:")
    db.init_db(conn)
    repo.upsert_class(conn, "6A5", 0)
    # Chỉ seed đúng 1 môn -- "Giáo dục thể chất" (map từ "GDTC") CHƯA tồn tại,
    # buộc importer phải tự tạo mới thay vì tìm thấy sẵn trong subj_name_to_id.
    repo.upsert_subject(conn, "Toán học", ROLE_THUONG, 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HKI_K6"
    ws.cell(4, 4, 1)       # hàng tuần: tuần 1 ở cột D
    ws.cell(4, 5, 2)       # tuần 2 ở cột E
    ws.cell(5, 2, "GDTC")  # Môn
    ws.cell(5, 4, 2)       # 2 tiết tuần 1
    ws.cell(5, 5, 2)       # 2 tiết tuần 2

    report = import_weekly_curriculum_from_excel(conn, wb)  # trước fix: NameError ở đây

    assert "Giáo dục thể chất" in report["subjects_mapped"]
    new_subj_id = repo.get_subject_by_name(conn, "Giáo dục thể chất")
    assert new_subj_id is not None
    subj = next(s for s in repo.list_subjects(conn) if s.subject_id == new_subj_id)
    assert subj.role_code == ROLE_THUONG
