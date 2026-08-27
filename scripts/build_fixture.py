"""Script to generate tests/fixtures/TKB_9lop_moi.xlsm and export_template.xlsm
with 8 classes, 16 subjects, 17 teachers, 4-3 frame, Week 1 (C) and Week 2 (L) quotas,
and all constraints from Phân công chuyên môn.xlsx and Định lượng số tiết theo tuần.
"""
from __future__ import annotations

import os
import sys
import shutil
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core import scheduler as sched
from data import db, repository as repo

FIXTURE_PATH = os.path.join(os.path.dirname(__file__), "..", "tests", "fixtures", "TKB_9lop_moi.xlsm")
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "io_excel", "export_template.xlsm")

def generate():
    # Make backup
    if os.path.exists(FIXTURE_PATH) and not os.path.exists(FIXTURE_PATH + ".bak"):
        shutil.copyfile(FIXTURE_PATH, FIXTURE_PATH + ".bak")

    # 1. First, setup DB in memory and run scheduler to get optimal conflict-free schedule for Parity C and L
    conn = db.get_connection(":memory:")
    db.init_db(conn)

    class_names = ["6A5", "6A6", "7A4", "7A5", "8A5", "8A6", "9A5", "9A6"]
    class_ids = {name: repo.upsert_class(conn, name, sort_order=i) for i, name in enumerate(class_names)}

    subject_defs = [
        ("Toán học", 3),
        ("Ngữ văn", 3),
        ("Ngoại ngữ", 1),
        ("Khoa học tự nhiên (Vật lý)", 0),
        ("Khoa học tự nhiên (Hoá học)", 0),
        ("Khoa học tự nhiên (Sinh học)", 0),
        ("Lịch sử và địa lý (Lịch sử)", 0),
        ("Lịch sử và địa lý (Địa lý)", 0),
        ("Giáo dục công dân", 0),
        ("Công nghệ", 0),
        ("Tin", 0),
        ("Giáo dục thể chất", 4),
        ("Nội dung giáo dục của địa phương", 0),
        ("Hoạt động trải nghiệm, hướng nghiệp", 5),
        ("Nghệ thuật (Nhạc)", 0),
        ("Nghệ thuật (Mỹ thuật)", 0),
    ]
    subject_ids = {name: repo.upsert_subject(conn, name, role_code=rc, sort_order=i) for i, (name, rc) in enumerate(subject_defs)}

    teacher_defs = [
        ("Hà", "Phó hiệu trưởng", True, False),
        ("Huyền Ly", "Tổ trưởng", True, False),
        ("Nguyễn Ly", "GVCN", True, True),
        ("Sơn", "Tổ phó", True, False),
        ("Hoà", "GVCN", True, True),
        ("Khu", "Thư ký HĐ", True, False),
        ("Hồng", "", False, False),
        ("Minh Anh", "GVCN", True, True),
        ("Lệ", "GVCN", True, True),
        ("Thành", "Tổ trưởng", True, False),
        ("Lan", "GVCN", True, True),
        ("Trung", "Tổ phó", True, False),
        ("Giang", "GVCN", True, True),
        ("Hoa", "GVCN", True, True),
        ("Lan Ly", "Hỗ trợ TPT", True, False),
        ("Uyên", "GVCN", True, True),
        ("Nhung", "", True, False),
    ]
    teacher_ids = {name: repo.upsert_teacher(conn, name, role=r, must_monday=mm, is_gvcn=ig) for name, r, mm, ig in teacher_defs}

    matrix = {
        "Toán học": {"6A5": "Huyền Ly", "6A6": "Huyền Ly", "7A4": "Lệ", "7A5": "Lệ", "8A5": "Minh Anh", "8A6": "Minh Anh", "9A5": "Nguyễn Ly", "9A6": "Nguyễn Ly"},
        "Ngữ văn": {"6A5": "Hoa", "6A6": "Hoa", "7A4": "Nhung", "7A5": "Thành", "8A5": "Uyên", "8A6": "Uyên", "9A5": "Lan", "9A6": "Lan"},
        "Ngoại ngữ": {"6A5": "Giang", "6A6": "Trung", "7A4": "Trung", "7A5": "Trung", "8A5": "Trung", "8A6": "Trung", "9A5": "Giang", "9A6": "Giang"},
        "Khoa học tự nhiên (Vật lý)": {"6A5": "Sơn", "6A6": "Sơn", "7A4": "Huyền Ly", "7A5": "Huyền Ly", "8A5": "Nguyễn Ly", "8A6": "Nguyễn Ly", "9A5": "Huyền Ly", "9A6": "Huyền Ly"},
        "Khoa học tự nhiên (Hoá học)": {"6A5": "Sơn", "6A6": "Sơn", "7A4": "Hoà", "7A5": "Hoà", "8A5": "Khu", "8A6": "Khu", "9A5": "Khu", "9A6": "Khu"},
        "Khoa học tự nhiên (Sinh học)": {"6A5": "Sơn", "6A6": "Sơn", "7A4": "Hoà", "7A5": "Hoà", "8A5": "Hoà", "8A6": "Hoà", "9A5": "Sơn", "9A6": "Sơn"},
        "Lịch sử và địa lý (Lịch sử)": {"6A5": "Hoa", "6A6": "Hoa", "7A4": "Thành", "7A5": "Thành", "8A5": "Thành", "8A6": "Thành", "9A5": "Thành", "9A6": "Thành"},
        "Lịch sử và địa lý (Địa lý)": {"6A5": "Nhung", "6A6": "Nhung", "7A4": "Nhung", "7A5": "Nhung", "8A5": "Nhung", "8A6": "Nhung", "9A5": "Nhung", "9A6": "Nhung"},
        "Giáo dục công dân": {"6A5": "Khu", "6A6": "Khu", "7A4": "Lan Ly", "7A5": "Lan Ly", "8A5": "Hoà", "8A6": "Uyên", "9A5": "Uyên", "9A6": "Uyên"},
        "Công nghệ": {"6A5": "Huyền Ly", "6A6": "Huyền Ly", "7A4": "Sơn", "7A5": "Sơn", "8A5": "Sơn", "8A6": "Sơn", "9A5": "Khu", "9A6": "Khu"},
        "Tin": {"6A5": "Minh Anh", "6A6": "Sơn", "7A4": "Khu", "7A5": "Khu", "8A5": "Lệ", "8A6": "Lệ", "9A5": "Minh Anh", "9A6": "Minh Anh"},
        "Giáo dục thể chất": {c: "Hồng" for c in class_names},
        "Nội dung giáo dục của địa phương": {"6A5": "Lan", "6A6": "Lan", "7A4": "Khu", "7A5": "Khu", "8A5": "Nhung", "8A6": "Nhung", "9A5": "Thành", "9A6": "Thành"},
        "Hoạt động trải nghiệm, hướng nghiệp": {"6A5": "Giang", "6A6": "Hoa", "7A4": "Lệ", "7A5": "Hoà", "8A5": "Minh Anh", "8A6": "Uyên", "9A5": "Nguyễn Ly", "9A6": "Lan"},
        "Nghệ thuật (Nhạc)": {"6A5": "Hà", "6A6": "Hà", "7A4": "Lan Ly", "7A5": "Lan Ly", "8A5": "Lan Ly", "8A6": "Lan Ly", "9A5": "Hà", "9A6": "Hà"},
        "Nghệ thuật (Mỹ thuật)": {c: "Lan Ly" for c in class_names}
    }
    for subj_name, cls_map in matrix.items():
        for cls_name, t_name in cls_map.items():
            repo.set_assignment(conn, subject_ids[subj_name], class_ids[cls_name], teacher_ids[t_name])

    for cid in class_ids.values():
        repo.set_frame_template(conn, cid, morning_periods=4, afternoon_periods=3, study_sunday=False, allow_saturday=False)

    grade_quotas = {
        "6": {
            "C": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 0, "Khoa học tự nhiên (Hoá học)": 4, "Khoa học tự nhiên (Sinh học)": 0, "Lịch sử và địa lý (Lịch sử)": 1, "Lịch sử và địa lý (Địa lý)": 2, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1},
            "L": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 2, "Khoa học tự nhiên (Hoá học)": 2, "Khoa học tự nhiên (Sinh học)": 0, "Lịch sử và địa lý (Lịch sử)": 1, "Lịch sử và địa lý (Địa lý)": 2, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1}
        },
        "7": {
            "C": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 1, "Khoa học tự nhiên (Hoá học)": 1, "Khoa học tự nhiên (Sinh học)": 2, "Lịch sử và địa lý (Lịch sử)": 2, "Lịch sử và địa lý (Địa lý)": 1, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1},
            "L": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 1, "Khoa học tự nhiên (Hoá học)": 1, "Khoa học tự nhiên (Sinh học)": 2, "Lịch sử và địa lý (Lịch sử)": 2, "Lịch sử và địa lý (Địa lý)": 1, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1}
        },
        "8": {
            "C": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 1, "Khoa học tự nhiên (Hoá học)": 3, "Khoa học tự nhiên (Sinh học)": 0, "Lịch sử và địa lý (Lịch sử)": 1, "Lịch sử và địa lý (Địa lý)": 2, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1},
            "L": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 2, "Khoa học tự nhiên (Hoá học)": 2, "Khoa học tự nhiên (Sinh học)": 0, "Lịch sử và địa lý (Lịch sử)": 1, "Lịch sử và địa lý (Địa lý)": 2, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1}
        },
        "9": {
            "C": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 0, "Khoa học tự nhiên (Hoá học)": 3, "Khoa học tự nhiên (Sinh học)": 1, "Lịch sử và địa lý (Lịch sử)": 2, "Lịch sử và địa lý (Địa lý)": 1, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1},
            "L": {"Toán học": 4, "Ngữ văn": 4, "Ngoại ngữ": 3, "Khoa học tự nhiên (Vật lý)": 3, "Khoa học tự nhiên (Hoá học)": 0, "Khoa học tự nhiên (Sinh học)": 1, "Lịch sử và địa lý (Lịch sử)": 2, "Lịch sử và địa lý (Địa lý)": 1, "Giáo dục công dân": 1, "Công nghệ": 1, "Tin": 1, "Giáo dục thể chất": 2, "Nội dung giáo dục của địa phương": 1, "Hoạt động trải nghiệm, hướng nghiệp": 3, "Nghệ thuật (Nhạc)": 1, "Nghệ thuật (Mỹ thuật)": 1}
        }
    }
    for cls_name, cid in class_ids.items():
        grade = cls_name[0]
        for par in ["C", "L"]:
            for subj_name, count in grade_quotas[grade][par].items():
                repo.set_periods_per_week(conn, subject_ids[subj_name], cid, par, count)

    repo.add_unavailability(conn, teacher_ids["Hồng"], "*", "S", "4")
    repo.add_unavailability(conn, teacher_ids["Hồng"], "*", "C", "1")

    # Thứ 2 đi hết tiết 1 (Chào cờ). 4 ngày còn lại (T3-T6), mỗi GV ưu tiên tối đa 3 ngày nghỉ tiết 1:
    ban_p1 = {
        "Huyền Ly": ["3", "4", "5"],
        "Nguyễn Ly": ["3", "4", "6"],
        "Sơn": ["3", "5", "6"],
        "Khu": ["3", "5"],
    }
    for tname, wds in ban_p1.items():
        for wd in wds:
            repo.add_unavailability(conn, teacher_ids[tname], wd, "S", "1")

    # Run scheduler for Parity C
    inp_c = repo.build_scheduling_input(conn, parity="C", seed=2026)
    res_c = sched.run(inp_c)
    assert res_c.success, f"Parity C failed: {res_c.failure_reason}"
    print(f"Parity C solved in {res_c.attempts_tried} attempts!")

    # Schedule mapping: (class_id, weekday, session, period) -> subject_name
    subj_name_by_id = {s.subject_id: s.name for s in inp_c.subjects}
    assigned_schedule_c = {}
    for slot in inp_c.slots:
        sid = res_c.assignment.get(slot.slot_id)
        if sid and sid != -1:
            assigned_schedule_c[(slot.class_id, slot.ts.weekday, slot.ts.session, slot.ts.period)] = subj_name_by_id[sid]
        else:
            assigned_schedule_c[(slot.class_id, slot.ts.weekday, slot.ts.session, slot.ts.period)] = ""

    # Load existing fixture workbook (keep VBA)
    wb = openpyxl.load_workbook(FIXTURE_PATH, keep_vba=True)

    # 1. Update PhanCong
    ws_pc = wb["PhanCong"]
    # Clear old data
    for r in range(2, ws_pc.max_row + 1):
        for c in range(1, ws_pc.max_column + 1):
            ws_pc.cell(r, c).value = None
    
    ws_pc.cell(2, 1).value = "Môn \\ Lớp"
    for j, cname in enumerate(class_names):
        ws_pc.cell(2, 2 + j).value = cname
    ws_pc.cell(2, 2 + len(class_names) + 1).value = "MÃ VAI TRÒ"

    for i, (sname, rcode) in enumerate(subject_defs):
        r = 3 + i
        ws_pc.cell(r, 1).value = sname
        for j, cname in enumerate(class_names):
            ws_pc.cell(r, 2 + j).value = matrix[sname][cname]
        ws_pc.cell(r, 2 + len(class_names) + 1).value = rcode

    # 2. Update SoTiet
    ws_st = wb["SoTiet"]
    for r in range(2, ws_st.max_row + 1):
        for c in range(1, ws_st.max_column + 1):
            ws_st.cell(r, c).value = None
    
    ws_st.cell(2, 1).value = "Môn \\ Lớp"
    for j, cname in enumerate(class_names):
        ws_st.cell(2, 2 + j).value = f"{cname} C"
    odd_start_col = 2 + len(class_names) + 1  # Col 11 (K)
    for j, cname in enumerate(class_names):
        ws_st.cell(2, odd_start_col + j).value = f"{cname} L"

    for i, (sname, _) in enumerate(subject_defs):
        r = 3 + i
        ws_st.cell(r, 1).value = sname
        for j, cname in enumerate(class_names):
            grade = cname[0]
            ws_st.cell(r, 2 + j).value = grade_quotas[grade]["C"][sname]
            ws_st.cell(r, odd_start_col + j).value = grade_quotas[grade]["L"][sname]

    # Row 19: TỔNG
    ws_st.cell(19, 1).value = "TỔNG"
    for j in range(len(class_names)):
        col_letter_c = openpyxl.utils.get_column_letter(2 + j)
        ws_st.cell(19, 2 + j).value = f"=SUM({col_letter_c}3:{col_letter_c}18)"
        col_letter_l = openpyxl.utils.get_column_letter(odd_start_col + j)
        ws_st.cell(19, odd_start_col + j).value = f"=SUM({col_letter_l}3:{col_letter_l}18)"

    # 3. Update DinhMuc_GV
    ws_dm = wb["DinhMuc_GV"]
    # Clear old teacher rows
    for r in range(3, ws_dm.max_row + 1):
        for c in range(1, 10):
            ws_dm.cell(r, c).value = None

    ws_dm.cell(1, 8).value = "Chuẩn:"
    ws_dm.cell(1, 9).value = 19
    ws_dm.cell(1, 11).value = "Chức vụ"
    ws_dm.cell(1, 12).value = "Giảm"
    ws_dm.cell(2, 11).value = "GVCN"
    ws_dm.cell(2, 12).value = 4
    ws_dm.cell(3, 11).value = "Tổ trưởng"
    ws_dm.cell(3, 12).value = 3
    ws_dm.cell(4, 11).value = "Tổ phó"
    ws_dm.cell(4, 12).value = 1
    ws_dm.cell(5, 11).value = "Tổng phụ trách"
    ws_dm.cell(5, 12).value = 8
    ws_dm.cell(6, 11).value = "Hỗ trợ TPT"
    ws_dm.cell(6, 12).value = 5
    ws_dm.cell(7, 11).value = "Phó hiệu trưởng"
    ws_dm.cell(7, 12).value = 15
    ws_dm.cell(8, 11).value = "Thư ký HĐ"
    ws_dm.cell(8, 12).value = 2

    for i, (tname, role, must_monday, is_gvcn) in enumerate(teacher_defs):
        r = 3 + i
        ws_dm.cell(r, 1).value = tname
        ws_dm.cell(r, 2).value = role if role else None
        ws_dm.cell(r, 3).value = f"=IFERROR(VLOOKUP(B{r},$K$2:$L$8,2,0),0)"
        ws_dm.cell(r, 4).value = f"=$I$1-C{r}"
        ws_dm.cell(r, 5).value = f"=SUMPRODUCT((PhanCong!$B$3:$I$18=A{r})*SoTiet!$B$3:$I$18)"
        ws_dm.cell(r, 6).value = f"=SUMPRODUCT((PhanCong!$B$3:$I$18=A{r})*SoTiet!$K$3:$R$18)"
        ws_dm.cell(r, 7).value = f"=E{r}-D{r}"
        ws_dm.cell(r, 8).value = 1 if must_monday else 0
        ws_dm.cell(r, 9).value = 1 if is_gvcn else 0

    # 4. Update GV_Ban
    ws_gb = wb["GV_Ban"]
    for r in range(3, ws_gb.max_row + 1):
        for c in range(1, ws_gb.max_column + 1):
            ws_gb.cell(r, c).value = None

    ban_rows = [
        ("Hồng", "*", "S", "4"),
        ("Hồng", "*", "C", "1"),
    ]
    for tname, wds in ban_p1.items():
        for wd in wds:
            ban_rows.append((tname, wd, "S", "1"))
    for i, (tname, wd, sess, per) in enumerate(ban_rows):
        r = 3 + i
        ws_gb.cell(r, 1).value = tname
        ws_gb.cell(r, 2).value = wd
        ws_gb.cell(r, 3).value = sess
        ws_gb.cell(r, 4).value = per

    # 5. Update Khung
    ws_khung = wb["Khung"]
    for r in range(2, ws_khung.max_row + 1):
        for c in range(1, ws_khung.max_column + 1):
            ws_khung.cell(r, c).value = None

    # 8 classes, each 7 rows (4 morning + 3 afternoon)
    row_idx = 2
    for cname in class_names:
        # Sáng 4 tiết
        for p in range(1, 5):
            ws_khung.cell(row_idx, 1).value = cname
            ws_khung.cell(row_idx, 2).value = "S"
            ws_khung.cell(row_idx, 3).value = p
            for wd in range(2, 7):  # Thứ 2..6
                ws_khung.cell(row_idx, wd + 2).value = "x"
            row_idx += 1
        # Chiều 3 tiết
        for p in range(1, 4):
            ws_khung.cell(row_idx, 1).value = cname
            ws_khung.cell(row_idx, 2).value = "C"
            ws_khung.cell(row_idx, 3).value = p
            for wd in range(2, 5):  # Thứ 2..4
                ws_khung.cell(row_idx, wd + 2).value = "x"
            row_idx += 1

    # 6. Update TKB_Nhap
    ws_nhap = wb["TKB_Nhap"]
    for r in range(2, ws_nhap.max_row + 1):
        for c in range(1, ws_nhap.max_column + 1):
            ws_nhap.cell(r, c).value = None

    row_idx = 2
    for cname in class_names:
        cid = class_ids[cname]
        for p in range(1, 5):
            ws_nhap.cell(row_idx, 1).value = cname
            ws_nhap.cell(row_idx, 2).value = "S"
            ws_nhap.cell(row_idx, 3).value = p
            for wd in range(2, 8):
                subj = assigned_schedule_c.get((cid, wd, "S", p), "")
                ws_nhap.cell(row_idx, wd + 2).value = subj if subj else None
            row_idx += 1
        for p in range(1, 4):
            ws_nhap.cell(row_idx, 1).value = cname
            ws_nhap.cell(row_idx, 2).value = "C"
            ws_nhap.cell(row_idx, 3).value = p
            for wd in range(2, 8):
                subj = assigned_schedule_c.get((cid, wd, "C", p), "")
                ws_nhap.cell(row_idx, wd + 2).value = subj if subj else None
            row_idx += 1

    total_data_rows = row_idx - 1  # 57

    # 7. Update TKB (formulas)
    ws_tkb = wb["TKB"]
    for r in range(2, ws_tkb.max_row + 1):
        for c in range(1, ws_tkb.max_column + 1):
            ws_tkb.cell(r, c).value = None

    for r in range(2, total_data_rows + 1):
        ws_tkb.cell(r, 1).value = f"=TKB_Nhap!A{r}"
        ws_tkb.cell(r, 2).value = f"=TKB_Nhap!B{r}"
        ws_tkb.cell(r, 3).value = f"=TKB_Nhap!C{r}"
        for col_idx, col_letter in enumerate(["D", "E", "F", "G", "H", "I"], start=4):
            ws_tkb.cell(r, col_idx).value = (
                f'=IF(TKB_Nhap!{col_letter}{r}="","",TKB_Nhap!{col_letter}{r}&CHAR(10)&'
                f'IF(T(INDEX(PhanCong!$B$3:$I$18,MATCH(TKB_Nhap!{col_letter}{r},PhanCong!$A$3:$A$18,0),'
                f'MATCH($A{r},PhanCong!$B$2:$I$2,0)))="","(chưa PC GV)","GV: "&'
                f'INDEX(PhanCong!$B$3:$I$18,MATCH(TKB_Nhap!{col_letter}{r},PhanCong!$A$3:$A$18,0),'
                f'MATCH($A{r},PhanCong!$B$2:$I$2,0))))'
            )

    # 8. Update TKB_GV (formulas)
    ws_gv = wb["TKB_GV"]
    for r in range(2, ws_gv.max_row + 1):
        for c in range(1, ws_gv.max_column + 1):
            ws_gv.cell(r, c).value = None

    for r in range(2, total_data_rows + 1):
        ws_gv.cell(r, 1).value = f"=TKB_Nhap!A{r}"
        ws_gv.cell(r, 2).value = f"=TKB_Nhap!B{r}"
        ws_gv.cell(r, 3).value = f"=TKB_Nhap!C{r}"
        for col_idx, col_letter in enumerate(["D", "E", "F", "G", "H", "I"], start=4):
            ws_gv.cell(r, col_idx).value = (
                f'=IF(TKB_Nhap!{col_letter}{r}="","",IFERROR(T(INDEX(PhanCong!$B$3:$I$18,'
                f'MATCH(TKB_Nhap!{col_letter}{r},PhanCong!$A$3:$A$18,0),'
                f'MATCH($A{r},PhanCong!$B$2:$I$2,0))),""))'
            )

    # 9. Update KiemTra (formulas)
    ws_kt = wb["KiemTra"]
    for r in range(2, ws_kt.max_row + 1):
        for c in range(1, ws_kt.max_column + 1):
            ws_kt.cell(r, c).value = None

    ws_kt.cell(3, 1).value = "Môn \\ Lớp"
    for j, cname in enumerate(class_names):
        ws_kt.cell(3, 2 + j).value = cname

    for i, (sname, _) in enumerate(subject_defs):
        r = 4 + i
        ws_kt.cell(r, 1).value = sname
        for j, cname in enumerate(class_names):
            col_letter = openpyxl.utils.get_column_letter(2 + j)
            ws_kt.cell(r, 2 + j).value = (
                f'=SUMPRODUCT((TKB_Nhap!$A$2:$A${total_data_rows}={col_letter}$3)*'
                f'(TKB_Nhap!$D$2:$J${total_data_rows}=$A{r}))-'
                f'IFERROR(N(IF(TuanConfig!$B$2="C",INDEX(SoTiet!$B$3:$I$18,MATCH($A{r},SoTiet!$A$3:$A$18,0),{j+1}),'
                f'INDEX(SoTiet!$K$3:$R$18,MATCH($A{r},SoTiet!$A$3:$A$18,0),{j+1}))),0)'
            )

    ws_kt.cell(20, 1).value = "TỔNG"
    for j in range(len(class_names)):
        col_letter = openpyxl.utils.get_column_letter(2 + j)
        ws_kt.cell(20, 2 + j).value = f"=SUM({col_letter}4:{col_letter}19)"

    # 10. Update TuanConfig
    ws_tc = wb["TuanConfig"]
    ws_tc.cell(1, 2).value = 2026
    ws_tc.cell(2, 2).value = "C"
    ws_tc.cell(4, 1).value = 1
    ws_tc.cell(4, 2).value = 2026
    ws_tc.cell(4, 3).value = "27/08/2026 22:00 [C]"
    ws_tc.cell(5, 1).value = 2
    ws_tc.cell(5, 2).value = 2026
    ws_tc.cell(5, 3).value = "27/08/2026 22:00 [L]"

    # Save to FIXTURE_PATH
    wb.save(FIXTURE_PATH)
    print("Saved updated fixture to:", FIXTURE_PATH)

    # Also save template (drop VBA or keep as is)
    wb_tpl = openpyxl.load_workbook(FIXTURE_PATH, keep_vba=True)
    wb_tpl.save(TEMPLATE_PATH)
    print("Saved updated template to:", TEMPLATE_PATH)

if __name__ == "__main__":
    generate()
