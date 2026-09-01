import openpyxl
import re
import sqlite3
import os
from copy import copy
from collections import defaultdict
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

def build_standard_tkb_ha():
    input_file = "TKB Hà.xlsx"
    output_file = "TKB_Ha_Chuan.xlsx"
    template_path = os.path.join("io_excel", "export_template.xlsm")
    
    # 1. Connect to DB to get metadata (roles, reduction, etc.)
    db_path = os.path.join("schools", "truong-thcs.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    classes = [
        {"class_id": 1, "name": "6A5", "sort_order": 0},
        {"class_id": 2, "name": "6A6", "sort_order": 1},
        {"class_id": 3, "name": "7A4", "sort_order": 2},
        {"class_id": 4, "name": "7A5", "sort_order": 3},
        {"class_id": 5, "name": "8A5", "sort_order": 4},
        {"class_id": 6, "name": "8A6", "sort_order": 5},
        {"class_id": 7, "name": "9A5", "sort_order": 6},
        {"class_id": 8, "name": "9A6", "sort_order": 7},
    ]
    class_names = [c["name"] for c in classes]

    subjects = [
        {"subject_id": 1, "name": "Toán học", "role_code": 1, "sort_order": 0},
        {"subject_id": 2, "name": "Ngữ văn", "role_code": 1, "sort_order": 1},
        {"subject_id": 3, "name": "Ngoại ngữ", "role_code": 1, "sort_order": 2},
        {"subject_id": 4, "name": "Khoa học tự nhiên (Vật lý)", "role_code": 0, "sort_order": 3},
        {"subject_id": 5, "name": "Khoa học tự nhiên (Hóa học)", "role_code": 0, "sort_order": 4},
        {"subject_id": 6, "name": "Khoa học tự nhiên (Sinh học)", "role_code": 0, "sort_order": 5},
        {"subject_id": 7, "name": "Lịch sử và Địa Lý (Lịch sử)", "role_code": 0, "sort_order": 6},
        {"subject_id": 8, "name": "Lịch sử và Địa Lý (Địa lý)", "role_code": 0, "sort_order": 7},
        {"subject_id": 9, "name": "GDCD", "role_code": 0, "sort_order": 8},
        {"subject_id": 10, "name": "Công nghệ", "role_code": 0, "sort_order": 9},
        {"subject_id": 11, "name": "Tin học", "role_code": 0, "sort_order": 10},
        {"subject_id": 12, "name": "Giáo dục thể chất", "role_code": 4, "sort_order": 11},
        {"subject_id": 13, "name": "Nội dung giáo dục của địa phương", "role_code": 0, "sort_order": 12},
        {"subject_id": 14, "name": "Hoạt động trải nghiệm, hướng nghiệp", "role_code": 5, "sort_order": 13},
        {"subject_id": 15, "name": "Nghệ thuật (Âm nhạc)", "role_code": 0, "sort_order": 14},
        {"subject_id": 16, "name": "Nghệ thuật (Mỹ thuật)", "role_code": 0, "sort_order": 15},
    ]

    teachers = [
        {"teacher_id": 1, "name": "Huyền Ly", "role": "", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 2, "name": "Lệ", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 3, "name": "Minh Anh", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 4, "name": "Nguyễn Ly", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 5, "name": "Hoa", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 6, "name": "Nhung", "role": "", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 7, "name": "Thành", "role": "", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 8, "name": "Uyên", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 9, "name": "Lan", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 10, "name": "Giang", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 11, "name": "Trung", "role": "", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 12, "name": "Sơn", "role": "", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 13, "name": "Hoà", "role": "GVCN", "must_monday": True, "is_gvcn": True},
        {"teacher_id": 14, "name": "Khu", "role": "Thư ký HĐ", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 15, "name": "Lan Ly", "role": "", "must_monday": True, "is_gvcn": False},
        {"teacher_id": 16, "name": "Hồng", "role": "", "must_monday": False, "is_gvcn": False},
        {"teacher_id": 17, "name": "Hà", "role": "Phó hiệu trưởng", "must_monday": True, "is_gvcn": False},
    ]

    role_reduction = {
        "GVCN": 4,
        "Tổ trưởng": 3,
        "Tổ phó": 1,
        "Tổng phụ trách": 8,
        "Phó hiệu trưởng": 15,
        "Hỗ trợ TPT": 5,
        "Thư ký HĐ": 2,
    }

    # 2. Parse TKB Hà.xlsx
    wb_ha = openpyxl.load_workbook(input_file, data_only=True)
    ws_ha = wb_ha["Sheet1"]

    teacher_map = {
        'Giang': 'Giang',
        'Hoa': 'Hoa',
        'N. Hoa': 'Hoa',
        'Lệ': 'Lệ',
        'Hòa': 'Hoà',
        'Hoà': 'Hoà',
        'M. Anh': 'Minh Anh',
        'Minh Anh': 'Minh Anh',
        'Uyên': 'Uyên',
        'N. Ly': 'Nguyễn Ly',
        'Nguyễn Ly': 'Nguyễn Ly',
        'H. Ly': 'Huyền Ly',
        'Huyền Ly': 'Huyền Ly',
        'L. Ly': 'Lan Ly',
        'Lan Ly': 'Lan Ly',
        'Lan': 'Lan',
        'Thành': 'Thành',
        'Trung': 'Trung',
        'Hồng': 'Hồng',
        'Khu': 'Khu',
        'Sơn': 'Sơn',
        'Nhung': 'Nhung',
        'Hà': 'Hà'
    }

    subject_map = {
        'HĐTN': 'Hoạt động trải nghiệm, hướng nghiệp',
        'Ngữ văn': 'Ngữ văn',
        'Toán': 'Toán học',
        'Toán học': 'Toán học',
        'Ngoại ngữ': 'Ngoại ngữ',
        'GDTC': 'Giáo dục thể chất',
        'LSĐL (Sử)': 'Lịch sử và Địa Lý (Lịch sử)',
        'LSĐL (Địa)': 'Lịch sử và Địa Lý (Địa lý)',
        'GDCD': 'GDCD',
        'Công nghệ': 'Công nghệ',
        'Tin học': 'Tin học',
        'GDĐP': 'Nội dung giáo dục của địa phương',
        'NT (ÂN)': 'Nghệ thuật (Âm nhạc)',
        'NT (MT)': 'Nghệ thuật (Mỹ thuật)',
        'KHTN (Hóa)': 'Khoa học tự nhiên (Hóa học)',
        'KHTN (Lý)': 'Khoa học tự nhiên (Vật lý)',
        'KHTN (Sinh)': 'Khoa học tự nhiên (Sinh học)',
        'KHTN': 'Khoa học tự nhiên (Hóa học)',
    }

    def parse_cell(val):
        if not val:
            return None, None
        val = str(val).strip()
        if val in ('.;', '.', ''):
            return None, None
        parts = re.split(r'\s*[-–_]\s*', val)
        if len(parts) == 2:
            subj_raw, tea_raw = parts[0].strip(), parts[1].strip()
        elif len(parts) == 1:
            subj_raw, tea_raw = parts[0].strip(), ''
        else:
            tea_raw = parts[-1].strip()
            subj_raw = '-'.join(parts[:-1]).strip()
        return subj_raw, tea_raw

    slot_defs = [
        (3, 'Thứ 2', 2, 'S', 1),
        (4, 'Thứ 2', 2, 'S', 2),
        (5, 'Thứ 2', 2, 'S', 3),
        (6, 'Thứ 2', 2, 'S', 4),
        (8, 'Thứ 2', 2, 'C', 1),
        (9, 'Thứ 2', 2, 'C', 2),
        (10, 'Thứ 2', 2, 'C', 3),
        (12, 'Thứ 3', 3, 'S', 1),
        (13, 'Thứ 3', 3, 'S', 2),
        (14, 'Thứ 3', 3, 'S', 3),
        (15, 'Thứ 3', 3, 'S', 4),
        (17, 'Thứ 3', 3, 'C', 1),
        (18, 'Thứ 3', 3, 'C', 2),
        (19, 'Thứ 3', 3, 'C', 3),
        (21, 'Thứ 4', 4, 'S', 1),
        (22, 'Thứ 4', 4, 'S', 2),
        (23, 'Thứ 4', 4, 'S', 3),
        (24, 'Thứ 4', 4, 'S', 4),
        (26, 'Thứ 4', 4, 'C', 1),
        (27, 'Thứ 4', 4, 'C', 2),
        (28, 'Thứ 4', 4, 'C', 3),
        (30, 'Thứ 5', 5, 'S', 1),
        (31, 'Thứ 5', 5, 'S', 2),
        (32, 'Thứ 5', 5, 'S', 3),
        (33, 'Thứ 5', 5, 'S', 4),
        (35, 'Thứ 6', 6, 'S', 1),
        (36, 'Thứ 6', 6, 'S', 2),
        (37, 'Thứ 6', 6, 'S', 3),
        (38, 'Thứ 6', 6, 'S', 4),
        (39, 'Thứ 6', 6, 'S', 5),
    ]

    tkb_grid = {} # (cls_name, wd, session, period) -> {"subject": s, "teacher": t}
    teacher_assignments = {} # (subj_name, cls_name) -> teacher_name
    period_counts_c = defaultdict(lambda: defaultdict(int)) # (subj_name, cls_name) -> count

    slot_teacher_classes = defaultdict(list)

    for r, day_name, wd, session, period in slot_defs:
        for c_idx, cls_name in enumerate(class_names):
            col = 3 + c_idx
            raw_val = ws_ha.cell(r, col).value
            if not raw_val or str(raw_val).strip() in ('', '.;'):
                continue
            subj_raw, tea_raw = parse_cell(raw_val)
            t_canonical = teacher_map.get(tea_raw, tea_raw)
            s_canonical = subject_map.get(subj_raw, subj_raw)

            tkb_grid[(cls_name, wd, session, period)] = {
                "subject": s_canonical,
                "teacher": t_canonical,
                "raw": raw_val
            }
            teacher_assignments[(s_canonical, cls_name)] = t_canonical
            period_counts_c[cls_name][s_canonical] += 1
            slot_teacher_classes[(t_canonical, wd, session, period)].append(cls_name)

    conflicts = {k: v for k, v in slot_teacher_classes.items() if len(v) > 1}

    # Load template
    wb = openpyxl.load_workbook(template_path)

    # 3. Helpers
    def _clear_values(ws, first_data_row: int = 2) -> None:
        for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    def _capture_row_style(ws, row_idx: int, n_cols: int) -> list:
        return [
            (copy(c.font), copy(c.fill), copy(c.border), copy(c.alignment), c.number_format)
            for c in (ws.cell(row_idx, col) for col in range(1, n_cols + 1))
        ]

    def _apply_row_style(ws, row_idx: int, style: list) -> None:
        for col, (font, fill, border, alignment, number_format) in enumerate(style, start=1):
            cell = ws.cell(row_idx, col)
            cell.font = font
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment
            cell.number_format = number_format

    def _detect_banding(ws, first_data_row: int, n_cols: int, max_scan: int = 200) -> tuple:
        style_a = _capture_row_style(ws, first_data_row, n_cols)
        for r in range(first_data_row + 1, min(first_data_row + max_scan, ws.max_row) + 1):
            fill = ws.cell(r, 1).fill
            if fill and fill.fill_type == "solid" and fill.fgColor.rgb not in (None, "00000000"):
                return style_a, _capture_row_style(ws, r, n_cols)
        return style_a, style_a

    def _autofit_sheet(ws, min_width: int = 8, max_width: int = 45, col_padding: int = 3,
                       row_height_per_line: float = 16, min_row_height: float = 18) -> None:
        col_widths: dict = {}
        row_lines: dict = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                lines = str(cell.value).split("\n")
                longest_line = max(len(line) for line in lines)
                col_widths[cell.column_letter] = max(col_widths.get(cell.column_letter, 0), longest_line)
                row_lines[cell.row] = max(row_lines.get(cell.row, 1), len(lines))
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = max(min_width, min(width + col_padding, max_width))
        for row_idx, n_lines in row_lines.items():
            ws.row_dimensions[row_idx].height = max(min_row_height, n_lines * row_height_per_line)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    weekdays = (2, 3, 4, 5, 6, 7)
    weekday_names = {2: "Thứ 2", 3: "Thứ 3", 4: "Thứ 4", 5: "Thứ 5", 6: "Thứ 6", 7: "Thứ 7", 8: "Chủ nhật"}

    # 4. Fill PhanCong
    ws_pc = wb["PhanCong"]
    _clear_values(ws_pc, first_data_row=2)
    n_classes = len(classes)
    n_subjects = len(subjects)
    code_col = 2 + n_classes + 1
    for i, cls in enumerate(classes):
        ws_pc.cell(2, 2 + i).value = cls["name"]
    ws_pc.cell(2, code_col).value = "MÃ VAI TRÒ"

    for r, subj in enumerate(subjects):
        row = 3 + r
        ws_pc.cell(row, 1).value = subj["name"]
        ws_pc.cell(row, code_col).value = subj["role_code"]
        for i, cls in enumerate(classes):
            tea = teacher_assignments.get((subj["name"], cls["name"]), "")
            ws_pc.cell(row, 2 + i).value = tea

    # 5. Fill SoTiet
    # Get DB periods_per_week for Odd weeks
    db_periods = {}
    for r in conn.execute('SELECT s.name as subj, c.name as cls, p.parity, p.periods FROM periods_per_week p JOIN subjects s ON p.subject_id = s.subject_id JOIN classes c ON p.class_id = c.class_id').fetchall():
        db_periods[(r['subj'], r['cls'], r['parity'])] = r['periods']

    ws_st = wb["SoTiet"]
    _clear_values(ws_st, first_data_row=2)
    odd_start_col = 2 + n_classes + 1
    for i, cls in enumerate(classes):
        ws_st.cell(2, 2 + i).value = f"{cls['name']} C"
        ws_st.cell(2, odd_start_col + i).value = f"{cls['name']} L"
    for r, subj in enumerate(subjects):
        row = 3 + r
        ws_st.cell(row, 1).value = subj["name"]
        for i, cls in enumerate(classes):
            # Even week: exact count from TKB Hà
            cnt_c = period_counts_c[cls["name"]].get(subj["name"], 0)
            cnt_l = db_periods.get((subj["name"], cls["name"], "L"), cnt_c)
            ws_st.cell(row, 2 + i).value = cnt_c
            ws_st.cell(row, odd_start_col + i).value = cnt_l

    # 6. Fill DinhMuc_GV
    ws_dm = wb["DinhMuc_GV"]
    _clear_values(ws_dm, first_data_row=2)
    ws_dm.cell(1, 8).value = "Chuẩn:"
    ws_dm.cell(1, 9).value = 19
    ws_dm.cell(1, 11).value = "Chức vụ"
    ws_dm.cell(1, 12).value = "Giảm"
    ws_dm.cell(1, 14).value = "Sàn tối thiểu:"
    ws_dm.cell(1, 15).value = 0
    ws_dm.cell(2, 1).value = "Tên GV"
    ws_dm.cell(2, 2).value = "Chức vụ"
    ws_dm.cell(2, 8).value = "Đi T2 (1/0)"
    ws_dm.cell(2, 9).value = "GVCN (1/0)"
    for r, t in enumerate(teachers):
        row = 3 + r
        ws_dm.cell(row, 1).value = t["name"]
        ws_dm.cell(row, 2).value = t["role"]
        ws_dm.cell(row, 8).value = int(t["must_monday"])
        ws_dm.cell(row, 9).value = int(t["is_gvcn"])
    for r, (role_name, reduction) in enumerate(role_reduction.items()):
        row = 2 + r
        ws_dm.cell(row, 11).value = role_name
        ws_dm.cell(row, 12).value = reduction

    # 7. Fill GV_Ban
    ws_gb = wb["GV_Ban"]
    _clear_values(ws_gb, first_data_row=2)
    ws_gb.cell(2, 1).value = "Giáo viên"
    ws_gb.cell(2, 2).value = "Thứ"
    ws_gb.cell(2, 3).value = "Buổi"
    ws_gb.cell(2, 4).value = "Tiết"
    unavailability_db = conn.execute('''
        SELECT t.name as tea, u.weekday, u.session, u.period
        FROM teacher_unavailability u
        JOIN teachers t ON u.teacher_id = t.teacher_id
        ORDER BY u.row_id
    ''').fetchall()
    for r, row_data in enumerate(unavailability_db):
        row = 3 + r
        ws_gb.cell(row, 1).value = row_data["tea"]
        ws_gb.cell(row, 2).value = row_data["weekday"]
        ws_gb.cell(row, 3).value = row_data["session"]
        ws_gb.cell(row, 4).value = row_data["period"]

    # 8. Define Frame for each class
    # 6A5, 6A6, 7A4, 7A5: S: 4, C: 3
    # 8A5, 8A6, 9A5, 9A6: S: 4 standard (short/outlier: T6 S=5), C: 3
    # Row layout: for classes 6-7: (S1..S4, C1..C3) -> 7 rows per class
    # for classes 8-9: (S1..S5, C1..C3) -> 8 rows per class
    class_sessions = {}
    for cls in classes:
        cls_name = cls["name"]
        if cls_name in ("6A5", "6A6", "7A4", "7A5"):
            class_sessions[cls_name] = [("S", 1), ("S", 2), ("S", 3), ("S", 4), ("C", 1), ("C", 2), ("C", 3)]
        else:
            class_sessions[cls_name] = [("S", 1), ("S", 2), ("S", 3), ("S", 4), ("S", 5), ("C", 1), ("C", 2), ("C", 3)]

    # 9. Fill Khung & TKB_Nhap
    ws_khung = wb["Khung"]
    ws_nh = wb["TKB_Nhap"]
    _clear_values(ws_khung, first_data_row=2)
    _clear_values(ws_nh, first_data_row=2)
    ws_nh.cell(1, 1).value = "LỚP HỌC"
    ws_nh.cell(1, 2).value = "BUỔI"
    ws_nh.cell(1, 3).value = "TIẾT THỨ"
    for i, wd in enumerate(weekdays):
        ws_nh.cell(1, 4 + i).value = weekday_names[wd]
    ws_nh.cell(1, 4 + len(weekdays)).value = weekday_names[8]

    row_idx = 2
    for cls in classes:
        cls_name = cls["name"]
        sessions = class_sessions[cls_name]
        for session, period in sessions:
            ws_nh.cell(row_idx, 1).value = cls_name
            ws_nh.cell(row_idx, 2).value = session
            ws_nh.cell(row_idx, 3).value = period
            for i, wd in enumerate(weekdays):
                col = 4 + i
                item = tkb_grid.get((cls_name, wd, session, period))
                if item:
                    ws_nh.cell(row_idx, col).value = item["subject"]
                    ws_khung.cell(row_idx, col).value = "x"
                else:
                    # Check if slot is active according to schedule rules (T2..T6 morning, T2..T4 afternoon)
                    if session == "S":
                        if period <= 4 and wd in (2, 3, 4, 5, 6):
                            ws_khung.cell(row_idx, col).value = "x"
                        elif period == 5 and wd == 6 and cls_name in ("8A5", "8A6", "9A5", "9A6"):
                            ws_khung.cell(row_idx, col).value = "x"
                    elif session == "C":
                        if period <= 3 and wd in (2, 3, 4):
                            ws_khung.cell(row_idx, col).value = "x"
            row_idx += 1

    # 10. Fill TKB, TKB_Mon, TKB_GV
    ws_tkb = wb["TKB"]
    ws_mon = wb["TKB_Nhap"] # We'll keep TKB_Nhap for import and copy or fill TKB_Mon if present
    # In template: we have TKB, TKB_GV, TKB_Nhap. Let's create TKB_Mon sheet as well!
    if "TKB_Mon" not in wb.sheetnames:
        ws_mon = wb.copy_worksheet(wb["TKB"])
        ws_mon.title = "TKB_Mon"
    else:
        ws_mon = wb["TKB_Mon"]
    ws_gv = wb["TKB_GV"]

    # Fill result sheets
    for ws_target, mode in [(ws_tkb, "full"), (ws_mon, "subject"), (ws_gv, "teacher")]:
        white_style, gray_style = _detect_banding(ws_target, first_data_row=2, n_cols=10)
        _clear_values(ws_target, first_data_row=2)
        ws_target.cell(1, 1).value = "LỚP HỌC"
        ws_target.cell(1, 2).value = "BUỔI"
        ws_target.cell(1, 3).value = "TIẾT THỨ"
        for i, wd in enumerate(weekdays):
            ws_target.cell(1, 4 + i).value = weekday_names[wd]
        ws_target.cell(1, 4 + len(weekdays)).value = weekday_names[8]

        row_idx = 2
        for cls_idx, cls in enumerate(classes):
            cls_name = cls["name"]
            band_style = white_style if cls_idx % 2 == 0 else gray_style
            sessions = class_sessions[cls_name]
            for session, period in sessions:
                _apply_row_style(ws_target, row_idx, band_style)
                ws_target.cell(row_idx, 1).value = cls_name
                ws_target.cell(row_idx, 2).value = session
                ws_target.cell(row_idx, 3).value = period
                for i, wd in enumerate(weekdays):
                    col = 4 + i
                    item = tkb_grid.get((cls_name, wd, session, period))
                    if item:
                        s_name = item["subject"]
                        t_name = item["teacher"]
                        if mode == "full":
                            ws_target.cell(row_idx, col).value = f"{s_name}\nGV: {t_name}"
                            ws_target.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="center")
                        elif mode == "subject":
                            ws_target.cell(row_idx, col).value = s_name
                            ws_target.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="center")
                        elif mode == "teacher":
                            ws_target.cell(row_idx, col).value = t_name
                            ws_target.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="center")
                            if (t_name, wd, session, period) in conflicts:
                                ws_target.cell(row_idx, col).fill = red_fill
                    else:
                        ws_target.cell(row_idx, col).value = ""
                row_idx += 1

    # 11. Fill TuanConfig
    ws_tc = wb["TuanConfig"]
    _clear_values(ws_tc, first_data_row=4)
    ws_tc.cell(1, 2).value = 2026
    ws_tc.cell(2, 2).value = "C"
    ws_tc.cell(4, 1).value = 1
    ws_tc.cell(4, 2).value = 2026
    ws_tc.cell(4, 3).value = "2026-09-07T00:00:00 [C]"

    # 12. Autofit all sheets
    for name in wb.sheetnames:
        _autofit_sheet(wb[name])

    # Reorder sheets logically:
    # 1. TKB (Full)
    # 2. TKB_Mon (Subject only)
    # 3. TKB_GV (Teacher only)
    # 4. PhanCong
    # 5. SoTiet
    # 6. DinhMuc_GV
    # 7. GV_Ban
    # 8. Khung
    # 9. TKB_Nhap
    # 10. TuanConfig
    # 11. HuongDan
    # 12. HuongDan_ChiTiet
    # 13. KiemTra
    sheet_order = ["TKB", "TKB_Mon", "TKB_GV", "PhanCong", "SoTiet", "DinhMuc_GV", "GV_Ban", "Khung", "TKB_Nhap", "TuanConfig", "KiemTra", "HuongDan", "HuongDan_ChiTiet"]
    
    # Filter only sheets that exist
    ordered_sheets = [s for s in sheet_order if s in wb.sheetnames]
    for s in wb.sheetnames:
        if s not in ordered_sheets:
            ordered_sheets.append(s)
    wb._sheets = [wb[s] for s in ordered_sheets]

    # Save
    wb.save(output_file)
    print(f"Successfully generated {output_file} with sheets: {wb.sheetnames}")

if __name__ == "__main__":
    build_standard_tkb_ha()
